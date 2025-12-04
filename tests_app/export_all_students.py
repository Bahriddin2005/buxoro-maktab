"""
Barcha o'quvchilarning test natijalarini Excel'ga export qilish
"""
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Max
from .models import TestAttempt, Test
from accounts.models import User

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def export_all_students_results(request):
    """Barcha o'quvchilarning test natijalarini Excel formatda export qilish"""
    
    # Faqat admin va o'qituvchilar
    if request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'openpyxl kutubxonasi o\'rnatilmagan'}, status=500)
    
    try:
        # Workbook yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Barcha Natijalar"
        
        # Title
        ws['A1'] = "BARCHA O'QUVCHILAR TEST NATIJALARI"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['№', 'Ism', 'Familiya', 'Sinf', 'Test nomi', 'Ball', 'Foiz', 'Baho']
        row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Barcha o'quvchilarni sinf bo'yicha olish
        students = User.objects.filter(role='student', is_verified=True).order_by('grade', 'last_name', 'first_name')
        
        if students.count() == 0:
            return JsonResponse({'error': 'O\'quvchilar topilmadi'}, status=404)
        
        # Ma'lumotlarni sinf bo'yicha guruhlash
        students_by_grade = {}
        for student in students:
            grade = student.grade or 'N/A'
            if grade not in students_by_grade:
                students_by_grade[grade] = []
            students_by_grade[grade].append(student)
        
        # Sinflarni tartiblab chiqarish
        grades = sorted(students_by_grade.keys(), key=lambda x: (x == 'N/A', int(x) if x != 'N/A' else 0))
        
        row = 4
        counter = 1
        
        # Har bir sinf uchun
        for grade in grades:
            grade_students = students_by_grade[grade]
            
            # SINF HEADER
            ws.cell(row=row, column=1, value=f"{grade}-SINF ({len(grade_students)} ta o'quvchi)")
            ws.merge_cells(f'A{row}:H{row}')
            header_cell = ws.cell(row=row, column=1)
            header_cell.font = Font(bold=True, size=14, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center")
            row += 1
            
            # Sinf ichidagi har bir o'quvchi
            for student in grade_students:
                # BARCHA tugallangan testlarni olish (faqat oxirgi emas, HAMMASI!)
                attempts = TestAttempt.objects.filter(
                    student=student,
                    is_completed=True
                ).select_related('test').order_by('test__subject', 'test__title', '-finished_at')
                
                if attempts.count() == 0:
                    # Hech qanday test yechmagan o'quvchi
                    ws.cell(row=row, column=1, value=counter)
                    ws.cell(row=row, column=2, value=student.first_name or student.username)
                    ws.cell(row=row, column=3, value=student.last_name or '')
                    ws.cell(row=row, column=4, value=student.grade or 'N/A')
                    ws.cell(row=row, column=5, value='Test yechmagan')
                    ws.cell(row=row, column=6, value='-')
                    ws.cell(row=row, column=7, value='-')
                    ws.cell(row=row, column=8, value='-')
                    
                    # Style
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    row += 1
                    counter += 1
                    continue
                
                # Har bir test natijasi uchun (HAMMASI!)
                for attempt in attempts:
                    percentage = attempt.percentage or 0
                    
                    # Baho'ni hisoblash
                    if percentage >= 81:
                        grade_text = "A'lo"
                        fill_color = "C6EFCE"  # Yashil
                    elif percentage >= 51:
                        grade_text = "Yaxshi"
                        fill_color = "FFEB9C"  # Sariq
                    elif percentage >= 31:
                        grade_text = "Qoniqarli"
                        fill_color = "BDD7EE"  # Ko'k
                    else:
                        grade_text = "Qoniqarsiz"
                        fill_color = "FFC7CE"  # Qizil
                    
                    # Ma'lumotlarni yozish
                    ws.cell(row=row, column=1, value=counter)
                    ws.cell(row=row, column=2, value=student.first_name or student.username)
                    ws.cell(row=row, column=3, value=student.last_name or '')
                    ws.cell(row=row, column=4, value=student.grade or 'N/A')
                    ws.cell(row=row, column=5, value=attempt.test.title)
                    ws.cell(row=row, column=6, value=f"{attempt.score}/{attempt.total_points}")
                    ws.cell(row=row, column=7, value=f"{percentage:.1f}%")
                    ws.cell(row=row, column=8, value=grade_text)
                    
                    # Style
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    
                    row += 1
                    counter += 1
            
            # Bo'sh qator sinf orasida
            row += 1
        
        # Umumiy statistika qo'shish
        row += 2
        ws[f'A{row}'] = "UMUMIY STATISTIKA"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        
        # Statistika hisoblash
        total_students = students.count()
        students_with_tests = students.filter(
            test_attempts__is_completed=True
        ).distinct().count()
        
        total_attempts = TestAttempt.objects.filter(
            student__role='student',
            is_completed=True
        ).count()
        
        avg_percentage = TestAttempt.objects.filter(
            student__role='student',
            is_completed=True
        ).aggregate(Avg('percentage'))['percentage__avg'] or 0
        
        ws[f'A{row}'] = f"Jami o'quvchilar: {total_students}"
        ws[f'C{row}'] = f"Test yechgan o'quvchilar: {students_with_tests}"
        ws[f'E{row}'] = f"Jami test natijalari: {total_attempts}"
        ws[f'G{row}'] = f"O'rtacha natija: {avg_percentage:.1f}%"
        
        # Ustun kengliklarini sozlash
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        
        # Response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="barcha_oquvchilar_natijalari.xlsx"'
        
        # Faylni saqlash
        wb.save(response)
        
        return response
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Export xatolik: {str(e)}'}, status=500)


# Import qo'shish
from django.db.models import Avg

