from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Count, Avg, Max, Min, Q
import json
import random
from .models import Test, Question, Choice, TestAttempt, Answer, TestResult, TestRetakeRequest
from accounts.models import User
from .views_overall import student_overall_results_view, student_export_results_view, test_api_view
from .export_all_students import export_all_students_results

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@login_required
def test_list_view(request):
    """List all available tests for students or created tests for teachers"""
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        # Pagination parametrlari
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 100))  # 100 ta test/sahifa
        
        # Filter parametrlari
        grade_filter = request.GET.get('grade', '')
        subject_filter = request.GET.get('subject', '')
        status_filter = request.GET.get('status', '')
        search_filter = request.GET.get('search', '')
        
        if request.user.role == 'student':
            tests = Test.objects.filter(
                is_active=True,
                grade=request.user.grade
            ).select_related('created_by').prefetch_related('questions').order_by('-created_at')
            
            # Qo'shimcha filterlar (o'quvchi o'z sinfidagi testlarni filtrlash uchun)
            if subject_filter:
                tests = tests.filter(subject=subject_filter)
            if search_filter:
                tests = tests.filter(
                    Q(title__icontains=search_filter) |
                    Q(description__icontains=search_filter)
                )
            
            # Bulk query optimizatsiyasi - N+1 query muammosini hal qilish
            test_ids = list(tests.values_list('id', flat=True))
            
            # Barcha attempt'larni bir marta olish
            attempts_dict = {}
            if test_ids:
                attempts = TestAttempt.objects.filter(
                    test_id__in=test_ids,
                    student=request.user
                ).values('test_id', 'id', 'is_completed', 'percentage')
                
                for att in attempts:
                    test_id = att['test_id']
                    if test_id not in attempts_dict:
                        attempts_dict[test_id] = {
                            'id': att['id'],
                            'is_completed': att['is_completed'],
                            'percentage': att.get('percentage')
                        }
            
            # Barcha retake request'larni bir marta olish
            retake_permissions = set()
            if test_ids:
                retakes = TestRetakeRequest.objects.filter(
                    student=request.user,
                    test_id__in=test_ids,
                    status='approved',
                    is_used=False
                ).values_list('test_id', flat=True)
                retake_permissions = set(retakes)
            
            test_data = []
            for test in tests:
                # Attempt ma'lumotlari
                attempt_data = attempts_dict.get(test.id)
                attempt = None
                if attempt_data:
                    class MinimalAttempt:
                        def __init__(self, data):
                            self.id = data['id']
                            self.is_completed = data['is_completed']
                            self.percentage = data.get('percentage')
                    attempt = MinimalAttempt(attempt_data)
                
                # Qayta ishlash ruxsati bormi tekshirish
                has_retake_permission = test.id in retake_permissions
                
                # O'quvchi testni yecha oladimi?
                can_attempt = test.is_active and (
                    attempt is None or  # Hali yechmagan
                    not attempt.is_completed or  # Davom ettirmoqda
                    has_retake_permission  # Qayta ishlash ruxsati bor
                )
                
                test_data.append({
                    'id': test.id,
                    'title': test.title,
                    'subject': test.subject,
                    'description': test.description,
                    'grade': test.grade,
                    'time_limit': test.time_limit,
                    'max_attempts': test.max_attempts,
                    'total_questions': test.total_questions,
                    'has_attempted': attempt is not None,
                    'attempt_score': round(attempt.percentage, 1) if attempt and attempt.is_completed and attempt.percentage else None,
                    'can_attempt': can_attempt,
                    'has_retake_permission': has_retake_permission,
                    'created_by': test.created_by.get_full_name() or test.created_by.username,
                    'created_at': test.created_at.isoformat(),
                    'start_time': test.start_time.isoformat() if test.start_time else None,
                    'end_time': test.end_time.isoformat() if test.end_time else None,
                })
            
            # Pagination
            start = (page - 1) * page_size
            end = start + page_size
            total_count = len(test_data)
            paginated_data = test_data[start:end]
            
            return JsonResponse({
                'tests': paginated_data,
                'user_role': 'student',
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size
                }
            })
            
        elif request.user.role == 'teacher':
            # Teacher sees ALL tests (not just their own) - LIMIT to first 100
            tests = Test.objects.all().select_related('created_by').order_by('-created_at')
            
            # Filterlar qo'shish
            if grade_filter:
                tests = tests.filter(grade=int(grade_filter))
            if subject_filter:
                tests = tests.filter(subject=subject_filter)
            if status_filter == 'active':
                tests = tests.filter(is_active=True)
            elif status_filter == 'inactive':
                tests = tests.filter(is_active=False)
            if search_filter:
                tests = tests.filter(
                    Q(title__icontains=search_filter) |
                    Q(description__icontains=search_filter)
                )
            
            # Get total count first
            total_count = tests.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            tests_page = tests[start:start + page_size]
            
            test_data = []
            for test in tests_page:
                attempt_count = TestAttempt.objects.filter(test=test, is_completed=True).count()
                test_data.append({
                    'id': test.id,
                    'title': test.title,
                    'subject': test.subject,
                    'description': test.description,
                    'grade': test.grade,
                    'total_questions': test.total_questions,
                    'is_active': test.is_active,
                    'created_at': test.created_at.isoformat(),
                    'created_by': test.created_by.get_full_name() or test.created_by.username,
                    'attempt_count': attempt_count,
                    'max_attempts': test.max_attempts,
                    'time_limit': test.time_limit,
                })
            
            return JsonResponse({
                'tests': test_data,
                'user_role': 'teacher',
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size
                }
            })
        
        elif request.user.role == 'admin':
            # Admin sees ALL tests - LIMIT to first 100
            tests = Test.objects.all().select_related('created_by').order_by('-created_at')
            
            # Filterlar qo'shish
            if grade_filter:
                tests = tests.filter(grade=int(grade_filter))
            if subject_filter:
                tests = tests.filter(subject=subject_filter)
            if status_filter == 'active':
                tests = tests.filter(is_active=True)
            elif status_filter == 'inactive':
                tests = tests.filter(is_active=False)
            if search_filter:
                tests = tests.filter(
                    Q(title__icontains=search_filter) |
                    Q(description__icontains=search_filter)
                )
            
            # Get total count first
            total_count = tests.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            tests_page = tests[start:start + page_size]
            
            test_data = []
            for test in tests_page:
                attempt_count = TestAttempt.objects.filter(test=test, is_completed=True).count()
                test_data.append({
                    'id': test.id,
                    'title': test.title,
                    'subject': test.subject,
                    'description': test.description,
                    'grade': test.grade,
                    'total_questions': test.total_questions,
                    'is_active': test.is_active,
                    'created_at': test.created_at.isoformat(),
                    'created_by': test.created_by.get_full_name() or test.created_by.username,
                    'attempt_count': attempt_count,
                    'max_attempts': test.max_attempts,
                    'time_limit': test.time_limit,
                })
            
            return JsonResponse({
                'tests': test_data,
                'user_role': 'admin',
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size
                }
            })
    
    return render(request, 'tests_app/test_list.html')

@login_required
@require_http_methods(["POST"])
def create_test(request):
    """Create a new test - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        required_fields = ['title', 'subject', 'grade', 'time_limit']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'error': f'{field} is required'}, status=400)
        
        with transaction.atomic():
            test = Test.objects.create(
                title=data['title'],
                description=data.get('description', ''),
                subject=data['subject'],
                grade=int(data['grade']),
                time_limit=int(data['time_limit']),
                created_by=request.user,
                start_time=data.get('start_time'),
                end_time=data.get('end_time'),
                max_attempts=data.get('max_attempts', 1),
                show_results=data.get('show_results', True),
                shuffle_questions=data.get('shuffle_questions', False)
            )
            
            questions_data = data.get('questions', [])
            for i, q_data in enumerate(questions_data):
                question = Question.objects.create(
                    test=test,
                    question_text=q_data['question_text'],
                    question_type=q_data['question_type'],
                    points=float(q_data.get('points', 1.0)),
                    order=i + 1,
                    explanation=q_data.get('explanation', '')
                )
                
                if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                    choices_data = q_data.get('choices', [])
                    for choice_data in choices_data:
                        Choice.objects.create(
                            question=question,
                            choice_text=choice_data['text'],
                            is_correct=choice_data.get('is_correct', False)
                        )
        
        return JsonResponse({
            'message': 'Test created successfully',
            'test_id': test.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def take_test_view(request, test_id):
    """Start or continue taking a test - Students only"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, is_active=True)
    
    if test.grade != request.user.grade:
        return JsonResponse({'error': 'This test is not for your grade'}, status=403)
    
    now = timezone.now()
    if test.start_time and now < test.start_time:
        return JsonResponse({'error': 'Test has not started yet'}, status=403)
    
    if test.end_time and now > test.end_time:
        return JsonResponse({'error': 'Test has ended'}, status=403)
    
    if request.method == 'POST':
        # correct_answers maydoni database'da yo'qligi uchun values() ishlatamiz
        existing_attempt_data = TestAttempt.objects.filter(
            test=test, 
            student=request.user
        ).values('id', 'is_completed').first()
        
        existing_attempt_id = None
        existing_attempt_completed = False
        if existing_attempt_data:
            existing_attempt_id = existing_attempt_data['id']
            existing_attempt_completed = existing_attempt_data['is_completed']
        
        # Agar test tugallangan bo'lsa, qayta ishlash ruxsati bormi tekshiramiz
        if existing_attempt_id and existing_attempt_completed:
            # Qayta ishlash ruxsati bormi?
            approved_retake = TestRetakeRequest.objects.filter(
                student=request.user,
                test=test,
                status='approved',
                is_used=False
            ).first()
            
            if not approved_retake:
                return JsonResponse({'error': 'Siz allaqachon bu testni topshirgansiz. Qayta topshirish uchun admin ruxsati kerak.'}, status=400)
            
            # Qayta ishlash ruxsati bor, yangi attempt yaratamiz (is_retake=True bilan)
            attempt = TestAttempt.objects.create(test=test, student=request.user, is_retake=True)
        elif not existing_attempt_id:
            # Birinchi marta test yechmoqda
            attempt = TestAttempt.objects.create(test=test, student=request.user)
        else:
            # Test tugallanmagan, davom ettirmoqda - to'liq obyektni yuklash
            attempt = TestAttempt.objects.filter(id=existing_attempt_id).only(
                'id', 'test', 'student', 'started_at', 'is_completed'
            ).first()
            if not attempt:
                # Agar topilmasa, yangi yaratamiz
                attempt = TestAttempt.objects.create(test=test, student=request.user)
        
        # Har bir o'quvchiga savollar random tartibda ko'rsatiladi
        # Query optimallashtirish - select_related va prefetch_related
        questions = list(test.questions.select_related().prefetch_related('choices').order_by('order'))
        
        # Shuffle questions for this student (har bir o'quvchi uchun boshqacha tartib)
        random.shuffle(questions)
        
        questions_data = []
        for question in questions:
            q_data = {
                'id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'points': question.points,
                'image': question.image.url if question.image and question.image.url else None
            }
            
            if question.question_type in ['single_choice', 'multiple_choice']:
                q_data['choices'] = [{
                    'id': choice.id,
                    'text': choice.choice_text
                } for choice in question.choices.all()]
            
            questions_data.append(q_data)
        
        return JsonResponse({
            'attempt_id': attempt.id,
            'questions': questions_data,
            'time_limit': test.time_limit,
            'started_at': attempt.started_at.isoformat(),
            'server_time': timezone.now().isoformat()  # Hozirgi server vaqti
        })
    
    # GET so'rovi - sahifa yuklash
    # Avval tugallangan testni tekshiramiz
    # correct_answers maydoni database'da yo'qligi uchun values() ishlatamiz
    existing_attempt_data = TestAttempt.objects.filter(
        test=test, 
        student=request.user
    ).values('id', 'is_completed').first()
    
    existing_attempt = None
    if existing_attempt_data:
        # Minimal object yaratish
        class MinimalAttempt:
            def __init__(self, data):
                self.id = data['id']
                self.is_completed = data['is_completed']
        
        existing_attempt = MinimalAttempt(existing_attempt_data)
    
    if existing_attempt and existing_attempt.is_completed:
        # Test allaqachon tugallangan
        # Qayta ishlash ruxsati bormi tekshiramiz
        approved_retake = TestRetakeRequest.objects.filter(
            student=request.user,
            test=test,
            status='approved',
            is_used=False
        ).first()
        
        if not approved_retake:
            # Ruxsat yo'q - natijalar sahifasiga yo'naltiramiz
            from django.contrib import messages
            messages.warning(request, 'Siz allaqachon bu testni topshirgansiz. Natijalarni ko\'ring yoki qayta topshirish uchun so\'rov yuboring.')
            return redirect('tests:test_results', test_id=test.id)
    
    return render(request, 'tests_app/take_test.html', {'test': test})

@login_required
@require_http_methods(["POST"])
def submit_answer(request, attempt_id):
    """Submit answer for a question"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        attempt = get_object_or_404(TestAttempt, id=attempt_id, student=request.user)
        
        if attempt.is_completed:
            return JsonResponse({'error': 'Test already completed'}, status=400)
        
        question_id = data.get('question_id')
        question = get_object_or_404(Question, id=question_id, test=attempt.test)
        
        answer, created = Answer.objects.get_or_create(
            attempt=attempt,
            question=question
        )
        
        # Clear previous answers
        answer.selected_choices.clear()
        answer.text_answer = ''
        
        if question.question_type == 'text_answer':
            answer.text_answer = data.get('text_answer', '')
        else:
            choice_ids = data.get('choice_ids', [])
            if choice_ids:
                choices = Choice.objects.filter(id__in=choice_ids, question=question)
                answer.selected_choices.set(choices)
        
        answer.save()
        
        # Update current question index for monitoring
        current_index = data.get('current_question_index', 0)
        if current_index is not None:
            attempt.current_question_index = current_index
            attempt.save(update_fields=['current_question_index'])
        
        # Verify answer was saved
        saved_choices = list(answer.selected_choices.values_list('id', flat=True))
        
        return JsonResponse({
            'message': 'Answer saved',
            'saved_choices': saved_choices,
            'text_answer': answer.text_answer
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def finish_test(request, attempt_id):
    """Finish the test and calculate score"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        attempt = get_object_or_404(TestAttempt, id=attempt_id, student=request.user)
        
        if attempt.is_completed:
            return JsonResponse({'error': 'Test already completed'}, status=400)
        
        attempt.finished_at = timezone.now()
        attempt.is_completed = True
        attempt.time_taken = attempt.finished_at - attempt.started_at
        
        # Calculate score using the method and get results
        results = attempt.calculate_score()
        
        # Use calculated values (do not access non-existing TestAttempt DB columns)
        correct_answers = results.get('correct_answers', 0)
        incorrect_answers = results.get('incorrect_answers', 0)
        unanswered = results.get('unanswered', 0)
        score = results.get('score', 0)
        total_points = results.get('total_points', 0)
        percentage = results.get('percentage', 0)

        # Update attempt fields (if model defines them they will be saved; otherwise these are harmless attributes)
        attempt.score = score
        attempt.total_points = total_points
        attempt.percentage = percentage
        
        test_result = TestResult.objects.create(
            attempt=attempt,
            correct_answers=correct_answers,
            incorrect_answers=incorrect_answers,
            unanswered=unanswered
        )
        test_result.grade = test_result.calculate_grade()
        test_result.save()
        
        attempt.save()
        
        # Agar bu qayta ishlash bo'lsa, retake request'ni ishlatilgan deb belgilaymiz
        if attempt.is_retake:
            retake_request = TestRetakeRequest.objects.filter(
                student=request.user,
                test=attempt.test,
                status='approved',
                is_used=False
            ).first()
            
            if retake_request:
                retake_request.is_used = True
                retake_request.save()
        
        # Create completion message
        total_questions = attempt.test.questions.count()
        answered_count = attempt.answers.count()
        all_answered = answered_count == total_questions
        
        completion_message = "Test yakunlandi!"
        if all_answered:
            completion_message = f"Ajoyib! Barcha {total_questions} ta savolga javob berdingiz!"
        else:
            completion_message = f"Test yakunlandi. {answered_count}/{total_questions} ta savolga javob berildi."
        
        # Grade message ni olish
        grade_message = test_result.get_grade_message() if test_result else ''
        
        # O'quvchining sinfidagi o'rinni hisoblash (ball bo'yicha)
        student_rank = None
        total_students_in_class = 0
        try:
            # Bir xil sinf va test bo'yicha barcha tugallangan attempt'larni olish
            # Har bir o'quvchining eng yaxshi natijasini olish
            from django.db.models import Max, Q
            # Har bir o'quvchi uchun eng yaxshi percentage va score ni topish
            best_attempts = TestAttempt.objects.filter(
                test=test,
                is_completed=True,
                student__grade=request.user.grade
            ).values('student_id').annotate(
                best_percentage=Max('percentage'),
                best_score=Max('score')
            ).order_by('-best_percentage', '-best_score')
            
            # O'quvchilarni list'ga o'tkazish
            attempts_list = list(best_attempts)
            total_students_in_class = len(attempts_list)
            
            # Joriy o'quvchining natijasini yangilash (agar yangi bo'lsa yoki yaxshiroq bo'lsa)
            current_student_found = False
            for attempt_data in attempts_list:
                if attempt_data['student_id'] == request.user.id:
                    current_student_found = True
                    # Agar joriy natija yaxshiroq bo'lsa, yangilash
                    if percentage > (attempt_data['best_percentage'] or 0):
                        attempt_data['best_percentage'] = percentage
                        attempt_data['best_score'] = score
                    break
            
            # Agar o'quvchi ro'yxatda yo'q bo'lsa, qo'shish
            if not current_student_found:
                attempts_list.append({
                    'student_id': request.user.id,
                    'best_percentage': percentage,
                    'best_score': score
                })
                total_students_in_class += 1
            
            # O'quvchilarni tartiblash (percentage, keyin score bo'yicha)
            attempts_list.sort(key=lambda x: (x.get('best_percentage') or 0, x.get('best_score') or 0), reverse=True)
            
            # O'quvchining o'rinni topish
            for rank, attempt_data in enumerate(attempts_list, 1):
                if attempt_data['student_id'] == request.user.id:
                    student_rank = rank
                    break
        except Exception as e:
            # Xatolik bo'lsa, o'rin hisoblanmaydi
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error calculating rank: {str(e)}")
        
        return JsonResponse({
            'message': completion_message,
            'results': {
                'score': results['score'],
                'total_points': results['total_points'],
                'percentage': results['percentage'],
                'grade': test_result.grade,
                'grade_message': grade_message,
                'correct_answers': correct_answers,
                'incorrect_answers': incorrect_answers,
                'unanswered': unanswered,
                'time_taken': str(attempt.time_taken),
                'all_answered': all_answered,
                'answered_count': answered_count,
                'total_questions': total_questions,
                'rank': student_rank,  # O'rin
                'total_students': total_students_in_class,  # Jami o'quvchilar soni
                'incorrect_questions': results.get('incorrect_questions', [])
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def test_results_view(request, test_id):
    """View test results - Teachers can see all, students see their own"""
    test = get_object_or_404(Test, id=test_id)
    
    if request.headers.get('Accept') == 'application/json':
        if request.user.role == 'student':
            if test.grade != request.user.grade:
                return JsonResponse({'error': 'Access denied'}, status=403)
            
            # correct_answers maydoni database'da yo'qligi uchun values() ishlatamiz
            attempt_data = TestAttempt.objects.filter(
                test=test, 
                student=request.user
            ).values('id', 'is_completed', 'percentage', 'score', 'total_points', 'finished_at').first()
            
            attempt = None
            attempt_id = None
            if attempt_data:
                attempt_id = attempt_data['id']
                # Minimal object yaratish
                class MinimalAttempt:
                    def __init__(self, data):
                        self.id = data['id']
                        self.is_completed = data['is_completed']
                        self.percentage = data.get('percentage')
                        self.score = data.get('score')
                        self.total_points = data.get('total_points')
                        self.finished_at = data.get('finished_at')
                
                attempt = MinimalAttempt(attempt_data)
            if not attempt or not attempt.is_completed:
                return JsonResponse({'error': 'Test not completed'}, status=404)
            
            # Get detailed results from TestResult
            from tests_app.models import TestResult
            test_result = TestResult.objects.filter(attempt_id=attempt_id).first()
            correct_answers = test_result.correct_answers if test_result else 0
            incorrect_answers = test_result.incorrect_answers if test_result else 0
            unanswered = test_result.unanswered if test_result else 0
            
            # Get incorrect questions details
            incorrect_questions = []
            unanswered_questions = []
            if test_result:
                # Get all answers for this attempt
                from tests_app.models import Answer
                answers = Answer.objects.filter(attempt_id=attempt_id).select_related('question')
                answered_question_ids = set(answers.values_list('question_id', flat=True))
                
                # Get all questions for this test
                all_questions = test.questions.all()
                
                for answer in answers:
                    if not answer.is_correct():
                        question = answer.question
                        student_answer = answer.get_student_answer_text()
                        correct_answer = question.get_correct_answer_text()
                        
                        incorrect_questions.append({
                            'question_text': question.question_text,
                            'student_answer': student_answer,
                            'correct_answer': correct_answer,
                            'explanation': question.explanation or ''
                        })
                
                # Find unanswered questions
                for question in all_questions:
                    if question.id not in answered_question_ids:
                        unanswered_questions.append({
                            'question_text': question.question_text,
                            'correct_answer': question.get_correct_answer_text(),
                            'explanation': question.explanation or ''
                        })
            
            # Get time_taken from TestAttempt
            attempt_with_time = TestAttempt.objects.filter(id=attempt_id).values('time_taken').first()
            time_taken = attempt_with_time.get('time_taken') if attempt_with_time else None
            
            # Grade message ni olish
            grade_message = ''
            if test_result:
                grade_message = test_result.get_grade_message()
            
            result_data = {
                'student': request.user.username,
                'score': attempt.score,
                'total_points': attempt.total_points,
                'percentage': attempt.percentage,
                'grade': test_result.grade if test_result else '',
                'grade_message': grade_message,
                'time_taken': str(time_taken) if time_taken else '',
                'finished_at': attempt.finished_at.isoformat() if attempt.finished_at else '',
                'correct_answers': correct_answers,
                'incorrect_answers': incorrect_answers,
                'unanswered': unanswered,
                'all_answered': (unanswered == 0),
                'answered_count': correct_answers + incorrect_answers,
                'total_questions': correct_answers + incorrect_answers + unanswered,
                # O'quvchilarga xato qilgan savollarni ko'rsatmaslik
                # 'incorrect_questions': incorrect_questions,
                # 'unanswered_questions': unanswered_questions
            }
            
            # Test info qo'shish
            test_info = {
                'title': test.title,
                'subject': test.subject,
                'grade': test.grade,
                'total_questions': test.total_questions,
                'time_limit': test.time_limit
            }
            
            return JsonResponse({
                'result': result_data,
                'test_info': test_info,
                'user_role': 'student'
            })
        
        elif (request.user.role == 'teacher' and test.created_by == request.user) or request.user.role == 'admin':
            # Faqat har bir o'quvchining oxirgi (eng so'nggi) natijasini olish
            from django.db.models import Max
            
            # Har bir o'quvchi uchun eng so'nggi attempt ID'sini topish
            latest_attempts = TestAttempt.objects.filter(
                test=test, 
                is_completed=True
            ).values('student').annotate(
                latest_attempt_id=Max('id')
            ).values_list('latest_attempt_id', flat=True)
            
            # Faqat oxirgi attempt'larni olish
            # correct_answers maydoni database'da yo'qligi uchun defer() ishlatamiz
            attempts = TestAttempt.objects.filter(
                id__in=latest_attempts
            ).select_related('student', 'result').defer(
                'correct_answers',
                'incorrect_answers',
                'unanswered'
            ).order_by('student__grade', 'student__class_name', 'student__first_name', 'student__last_name')
            
            results_data = []
            for attempt in attempts:
                # Admin va o'qituvchilar uchun xato qilgan savollarni ham qo'shamiz
                incorrect_questions = []
                unanswered_questions = []
                
                if hasattr(attempt, 'result') and attempt.result:
                    # Get all answers for this attempt
                    answers = Answer.objects.filter(attempt=attempt).select_related('question')
                    answered_question_ids = set(answers.values_list('question_id', flat=True))
                    
                    # Get all questions for this test
                    all_questions = test.questions.all()
                    
                    for answer in answers:
                        if not answer.is_correct():
                            question = answer.question
                            student_answer = answer.get_student_answer_text()
                            correct_answer = question.get_correct_answer_text()
                            
                            incorrect_questions.append({
                                'question_text': question.question_text,
                                'student_answer': student_answer,
                                'correct_answer': correct_answer,
                                'explanation': question.explanation or ''
                            })
                    
                    # Find unanswered questions
                    for question in all_questions:
                        if question.id not in answered_question_ids:
                            unanswered_questions.append({
                                'question_text': question.question_text,
                                'correct_answer': question.get_correct_answer_text(),
                                'explanation': question.explanation or ''
                            })
                
                results_data.append({
                    'student': {
                        'username': attempt.student.username,
                        'first_name': attempt.student.first_name,
                        'last_name': attempt.student.last_name,
                        'student_id': attempt.student.student_id,
                        'class_name': attempt.student.class_name,
                        'grade': attempt.student.grade
                    },
                    'score': attempt.score,
                    'total_points': attempt.total_points,
                    'percentage': attempt.percentage,
                    'grade': attempt.result.grade if hasattr(attempt, 'result') and attempt.result else '',
                    'time_taken': str(attempt.time_taken) if attempt.time_taken else '',
                    'finished_at': attempt.finished_at.isoformat() if attempt.finished_at else '',
                    'correct_answers': attempt.result.correct_answers if hasattr(attempt, 'result') and attempt.result else 0,
                    'incorrect_answers': attempt.result.incorrect_answers if hasattr(attempt, 'result') and attempt.result else 0,
                    'unanswered': attempt.result.unanswered if hasattr(attempt, 'result') and attempt.result else 0,
                    # Admin va o'qituvchilar uchun xato qilgan savollarni ko'rsatish
                    'incorrect_questions': incorrect_questions,
                    'unanswered_questions': unanswered_questions
                })
            
            # Test info qo'shish
            test_info = {
                'title': test.title,
                'subject': test.subject,
                'grade': test.grade,
                'total_questions': test.total_questions,
                'time_limit': test.time_limit
            }
            
            return JsonResponse({
                'results': results_data,
                'test_info': test_info,
                'user_role': request.user.role
            })
        
        else:
            return JsonResponse({'error': 'Access denied'}, status=403)
    
    return render(request, 'tests_app/test_results.html', {
        'test': test,
        'user_role': request.user.role
    })

@login_required
def export_results(request, test_id):
    """Export test results to Excel - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    attempts = TestAttempt.objects.filter(test=test, is_completed=True).select_related('student', 'result').order_by('student__grade', 'student__class_name', 'student__first_name', 'student__last_name')
    
    
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Header qo'shish
    headers = [
        'Student Username', 'First Name', 'Last Name', 'Student ID', 'Grade', 
        'Class', 'Score', 'Total Points', 'Percentage', 'Grade Result',
        'Correct Answers', 'Incorrect Answers', 'Unanswered', 'Time Taken', 'Finished At'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ma'lumotlarni qo'shish
    for row, attempt in enumerate(attempts, 2):
        data = [
            attempt.student.username,
            attempt.student.first_name,
            attempt.student.last_name,
            attempt.student.student_id or '',
            attempt.student.grade or '',
            attempt.student.class_name or '',
            attempt.score,
            attempt.total_points,
            attempt.percentage,
            attempt.result.grade if hasattr(attempt, 'result') and attempt.result else '',
            attempt.result.correct_answers if hasattr(attempt, 'result') and attempt.result else 0,
            attempt.result.incorrect_answers if hasattr(attempt, 'result') and attempt.result else 0,
            attempt.result.unanswered if hasattr(attempt, 'result') and attempt.result else 0,
            str(attempt.time_taken),
            attempt.finished_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Excel faylini qaytarish
    from django.http import HttpResponse
    from io import BytesIO
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{test.title}_results.xlsx"'
    
    return response

@login_required
def upload_questions(request, test_id):
    """Upload questions from Excel file - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({'error': 'No file uploaded'}, status=400)
            
            # Excel faylini openpyxl bilan o'qish
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Header qatorini o'qish
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value.lower().replace(' ', '_'))
            
            # Kerakli ustunlarni tekshirish
            required_columns = ['question_text', 'question_type', 'points']
            for col in required_columns:
                if col not in headers:
                    return JsonResponse({'error': f'Missing column: {col}'}, status=400)
            
            questions_created = 0
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                    if not row[0]:  # question_text bo'sh bo'lsa
                        continue
                    
                    row_data = dict(zip(headers, row))
                    
                    question = Question.objects.create(
                        test=test,
                        question_text=row_data['question_text'],
                        question_type=row_data['question_type'],
                        points=float(row_data.get('points', 1.0)),
                        order=row_num,
                        explanation=row_data.get('explanation', '')
                    )
                    
                    # Javob variantlarini qo'shish
                    if row_data['question_type'] in ['single_choice', 'multiple_choice']:
                        for i in range(1, 6):  # 5 tagacha variant
                            choice_key = f'choice_{i}'
                            correct_key = f'choice_{i}_correct'
                            
                            if choice_key in row_data and row_data[choice_key]:
                                is_correct = bool(row_data.get(correct_key, False))
                                Choice.objects.create(
                                    question=question,
                                    choice_text=row_data[choice_key],
                                    is_correct=is_correct
                                )
                    
                    questions_created += 1
            
            return JsonResponse({'message': f'{questions_created} questions uploaded successfully'})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'tests_app/upload_questions.html', {'test': test})

@login_required
def create_test_view(request):
    """Create new test - only for teachers"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'GET':
        return render(request, 'tests_app/create_test.html')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Test yaratish
                test = Test.objects.create(
                    title=request.POST.get('title'),
                    description=request.POST.get('description', ''),
                    subject=request.POST.get('subject'),
                    grade=int(request.POST.get('grade')),
                    time_limit=int(request.POST.get('time_limit', 45)),
                    max_attempts=int(request.POST.get('max_attempts', 1)),
                    show_results=bool(request.POST.get('show_results')),
                    is_active=bool(request.POST.get('is_active')),
                    created_by=request.user
                )
                
                # Savollar qo'shish
                question_texts = request.POST.getlist('question_text[]')
                question_types = request.POST.getlist('question_type[]')
                points_list = request.POST.getlist('points[]')
                explanations = request.POST.getlist('explanation[]')
                question_images = request.FILES.getlist('question_image[]')
                
                for i, question_text in enumerate(question_texts):
                    if not question_text.strip():
                        continue
                    
                    # Rasm faylini tekshirish va yuklash
                    question_image = None
                    if i < len(question_images):
                        uploaded_file = question_images[i]
                        # Fayl mavjudligini va hajmini tekshirish
                        if uploaded_file and uploaded_file.size > 0:
                            # Maksimal hajm: 1GB
                            MAX_FILE_SIZE = 1024 * 1024 * 1024  # 1GB
                            if uploaded_file.size > MAX_FILE_SIZE:
                                return JsonResponse({
                                    'success': False, 
                                    'error': f'Savol {i+1} uchun rasm hajmi 1GB dan oshmasligi kerak!'
                                }, status=400)
                            # Fayl turini tekshirish
                            if not uploaded_file.content_type.startswith('image/'):
                                return JsonResponse({
                                    'success': False, 
                                    'error': f'Savol {i+1} uchun faqat rasm fayllari ruxsat etilgan!'
                                }, status=400)
                            question_image = uploaded_file
                    
                    question = Question.objects.create(
                        test=test,
                        question_text=question_text,
                        question_type=question_types[i],
                        points=float(points_list[i]) if points_list[i] else 1.0,
                        order=i + 1,
                        explanation=explanations[i] if i < len(explanations) else '',
                        image=question_image
                    )
                    
                    # Javob variantlarini qo'shish
                    if question_types[i] != 'text_answer':
                        choices_key = f'choices_{i+1}[]'
                        correct_key = f'correct_choice_{i+1}'
                        
                        choices = request.POST.getlist(choices_key)
                        correct_index = request.POST.get(correct_key)
                        
                        for j, choice_text in enumerate(choices):
                            if choice_text.strip():
                                is_correct = str(j) == correct_index
                                Choice.objects.create(
                                    question=question,
                                    choice_text=choice_text,
                                    is_correct=is_correct
                                )
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Test muvaffaqiyatli yaratildi!',
                    'test_id': test.id
                })
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def test_info_view(request, test_id):
    """Get test information for display purposes"""
    test = get_object_or_404(Test, id=test_id)
    
    # Check access permissions
    if request.user.role == 'student' and test.grade != request.user.grade:
        return JsonResponse({'error': 'Access denied'}, status=403)
    elif request.user.role == 'teacher' and test.created_by != request.user:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    return JsonResponse({
        'title': test.title,
        'description': test.description,
        'subject': test.subject,
        'grade': test.grade,
        'time_limit': test.time_limit,
        'max_attempts': test.max_attempts,
        'total_questions': test.total_questions,
        'created_by': test.created_by.get_full_name() or test.created_by.username,
        'created_at': test.created_at.isoformat(),
        'start_time': test.start_time.isoformat() if test.start_time else None,
        'end_time': test.end_time.isoformat() if test.end_time else None,
    })

# all_results_view funksiyasi olib tashlandi - /tests/all-results/ bo'limi kerak emas edi

@login_required
@require_http_methods(["POST"])
def request_retake_view(request, test_id):
    """O'quvchi qayta ishlash so'rovi yuborish"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Faqat o\'quvchilar qayta ishlash so\'rovi yuborishi mumkin'}, status=403)
    
    test = get_object_or_404(Test, id=test_id)
    
    try:
        # O'quvchining oxirgi attempt'ini topamiz
        attempt = TestAttempt.objects.filter(test=test, student=request.user, is_completed=True).last()
        if not attempt:
            return JsonResponse({'error': 'Siz hali bu testni topshirmadingiz'}, status=400)
        
        # Qayta ishlash so'rashi mumkinmi?
        if not attempt.can_request_retake():
            return JsonResponse({'error': 'Allaqachon qayta ishlash so\'rovi yuborilgan yoki kutilmoqda'}, status=400)
        
        # JSON ma'lumotni olish
        try:
            if request.body:
                data = json.loads(request.body)
                reason = data.get('reason', '').strip()
            else:
                reason = ''
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Noto\'g\'ri JSON ma\'lumot. Iltimos, to\'g\'ri format kiriting.'}, status=400)
        
        # Agar sabab bo'sh bo'lsa, default sabab
        if not reason or len(reason) < 10:
            return JsonResponse({'error': 'Qayta ishlash sababi kamida 10 ta belgidan iborat bo\'lishi kerak'}, status=400)
        
        # Qayta ishlash so'rovini yaratamiz
        retake_request = TestRetakeRequest.objects.create(
            student=request.user,
            test=test,
            previous_attempt=attempt,
            reason=reason
        )
        
        return JsonResponse({
            'message': 'Qayta ishlash so\'rovi muvaffaqiyatli yuborildi!',
            'request_id': retake_request.id
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Xatolik yuz berdi: {str(e)}'}, status=500)

@login_required  
def retake_requests_view(request):
    """Admin qayta ishlash so'rovlarini ko'rish va boshqarish"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Faqat adminlar kirishi mumkin'}, status=403, json_dumps_params={'ensure_ascii': False})
    
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        try:
            # JSON API so'rovi
            status_filter = request.GET.get('status', 'all')
            
            # Raw SQL ishlatish - database'da maydon mavjudligini tekshirmasdan
            from django.db import connection
            from django.conf import settings
            
            db_engine = settings.DATABASES['default']['ENGINE']
            param_style = '?' if 'sqlite' in db_engine.lower() else '%s'
            
            # SQL query - faqat mavjud maydonlarni olish
            base_query = """
                SELECT 
                    trr.id as request_id,
                    trr.reason,
                    trr.status,
                    trr.admin_response,
                    trr.created_at,
                    trr.updated_at,
                    u.id as student_id,
                    u.username as student_username,
                    u.first_name as student_first_name,
                    u.last_name as student_last_name,
                    u.grade as student_grade,
                    u.class_name as student_class_name,
                    t.id as test_id,
                    t.title as test_title,
                    t.subject as test_subject,
                    ta.id as attempt_id,
                    ta.score as previous_score,
                    ta.percentage as previous_percentage,
                    approved.id as approved_by_id,
                    approved.first_name as approved_by_first_name,
                    approved.last_name as approved_by_last_name,
                    approved.username as approved_by_username
                FROM tests_app_testretakerequest trr
                INNER JOIN accounts_user u ON trr.student_id = u.id
                INNER JOIN tests_app_test t ON trr.test_id = t.id
                INNER JOIN tests_app_testattempt ta ON trr.previous_attempt_id = ta.id
                LEFT JOIN accounts_user approved ON trr.approved_by_id = approved.id
                WHERE 1=1
            """
            params = []
            
            if status_filter != 'all':
                base_query += " AND trr.status = " + param_style
                params.append(status_filter)
            
            base_query += " ORDER BY trr.created_at DESC"
            
            requests_data = []
            try:
                with connection.cursor() as cursor:
                    cursor.execute(base_query, params)
                    columns = [col[0] for col in cursor.description]
                    
                    for row in cursor.fetchall():
                        row_dict = dict(zip(columns, row))
                        
                        # Student ma'lumotlari
                        student_name = f"{row_dict.get('student_first_name') or ''} {row_dict.get('student_last_name') or ''}".strip()
                        if not student_name:
                            student_name = row_dict.get('student_username') or ''
                        
                        # Approved by
                        approved_by_name = None
                        if row_dict.get('approved_by_id'):
                            approved_by_name = f"{row_dict.get('approved_by_first_name') or ''} {row_dict.get('approved_by_last_name') or ''}".strip()
                            if not approved_by_name:
                                approved_by_name = row_dict.get('approved_by_username') or ''
                        
                        # Status display
                        status_display_map = {
                            'pending': 'Kutilmoqda',
                            'approved': 'Tasdiqlangan',
                            'rejected': 'Rad etilgan'
                        }
                        status_display = status_display_map.get(row_dict.get('status'), row_dict.get('status', ''))
                        
                        # Previous attempt ma'lumotlari
                        previous_score = row_dict.get('previous_score') or 0
                        previous_percentage = row_dict.get('previous_percentage') or 0
                        
                        requests_data.append({
                            'id': row_dict.get('request_id'),
                            'student_name': student_name,
                            'student_username': row_dict.get('student_username') or '',
                            'student_grade': row_dict.get('student_grade'),
                            'student_class': row_dict.get('student_class_name') or '',
                            'test_title': row_dict.get('test_title') or '',
                            'test_subject': row_dict.get('test_subject') or '',
                            'previous_score': previous_score,
                            'previous_percentage': previous_percentage,
                            'reason': row_dict.get('reason') or '',
                            'status': row_dict.get('status') or 'pending',
                            'status_display': status_display,
                            'admin_response': row_dict.get('admin_response') or '',
                            'approved_by': approved_by_name,
                            'created_at': row_dict.get('created_at').isoformat() if row_dict.get('created_at') else '',
                            'updated_at': row_dict.get('updated_at').isoformat() if row_dict.get('updated_at') else ''
                        })
            except Exception as query_error:
                # Fallback - Django ORM ishlatish (agar SQL xatolik bersa)
                requests_qs = TestRetakeRequest.objects.select_related(
                    'student', 'test', 'approved_by'
                ).order_by('-created_at')
                
                if status_filter != 'all':
                    requests_qs = requests_qs.filter(status=status_filter)
                
                for req in requests_qs:
                    try:
                        # Student ma'lumotlari
                        student_name = req.student.get_full_name() if req.student else ''
                        if not student_name:
                            student_name = f"{req.student.first_name or ''} {req.student.last_name or ''}".strip() or req.student.username if req.student else ''
                        
                        # Previous attempt ma'lumotlari - xavfsiz olish
                        previous_score = 0
                        previous_percentage = 0
                        if req.previous_attempt_id:
                            try:
                                # Direct values query
                                attempt_values = TestAttempt.objects.filter(
                                    id=req.previous_attempt_id
                                ).values('score', 'percentage').first()
                                if attempt_values:
                                    previous_score = attempt_values.get('score') or 0
                                    previous_percentage = attempt_values.get('percentage') or 0
                            except Exception:
                                previous_score = 0
                                previous_percentage = 0
                        
                        # Status display
                        status_display_map = {
                            'pending': 'Kutilmoqda',
                            'approved': 'Tasdiqlangan',
                            'rejected': 'Rad etilgan'
                        }
                        status_display = status_display_map.get(req.status, req.status)
                        
                        # Approved by
                        approved_by_name = None
                        if req.approved_by:
                            approved_by_name = req.approved_by.get_full_name()
                            if not approved_by_name:
                                approved_by_name = f"{req.approved_by.first_name or ''} {req.approved_by.last_name or ''}".strip() or req.approved_by.username
                        
                        requests_data.append({
                            'id': req.id,
                            'student_name': student_name,
                            'student_username': req.student.username if req.student else '',
                            'student_grade': req.student.grade if req.student and req.student.grade else None,
                            'student_class': req.student.class_name if req.student else '',
                            'test_title': req.test.title if req.test else '',
                            'test_subject': req.test.subject if req.test else '',
                            'previous_score': previous_score,
                            'previous_percentage': previous_percentage,
                            'reason': req.reason or '',
                            'status': req.status,
                            'status_display': status_display,
                            'admin_response': req.admin_response or '',
                            'approved_by': approved_by_name,
                            'created_at': req.created_at.isoformat() if req.created_at else '',
                            'updated_at': req.updated_at.isoformat() if req.updated_at else ''
                        })
                    except Exception:
                        continue
            
            return JsonResponse({
                'requests': requests_data,
                'total_count': len(requests_data)
            }, json_dumps_params={'ensure_ascii': False})
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'error': f'Ma\'lumotlarni yuklashda xatolik: {str(e)}',
                'requests': [],
                'total_count': 0
            }, status=500, json_dumps_params={'ensure_ascii': False})
    
    # HTML template
    return render(request, 'tests_app/retake_requests.html')

@login_required
@require_http_methods(["POST"])
def handle_retake_request_view(request, request_id):
    """Admin qayta ishlash so'rovini tasdiqlash yoki rad etish"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Faqat adminlar kirishi mumkin'}, status=403)
    
    retake_request = get_object_or_404(TestRetakeRequest, id=request_id)
    
    if retake_request.status != 'pending':
        return JsonResponse({'error': 'Bu so\'rov allaqachon ko\'rib chiqilgan'}, status=400)
    
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'approve' yoki 'reject'
        admin_response = data.get('admin_response', '').strip()
        
        if action not in ['approve', 'reject']:
            return JsonResponse({'error': 'Noto\'g\'ri harakat'}, status=400)
        
        retake_request.status = 'approved' if action == 'approve' else 'rejected'
        retake_request.admin_response = admin_response
        retake_request.approved_by = request.user
        retake_request.save()
        
        return JsonResponse({
            'message': f'So\'rov {"tasdiqlandi" if action == "approve" else "rad etildi"}!',
            'status': retake_request.status
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Noto\'g\'ri JSON ma\'lumot'}, status=400)
    except Exception as e:
        return JsonResponse({'error': 'Xatolik yuz berdi'}, status=500)


@login_required
def open_test_for_student(request, test_id, student_id):
    """Admin tomonidan o'quvchi uchun testni qayta ochish"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Ruxsat berilmagan'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST so\'rov talab qilinadi'}, status=405)
    
    try:
        test = Test.objects.get(id=test_id)
        student = User.objects.get(id=student_id, role='student')
        
        # O'quvchining bu testdagi avvalgi urinishlarini tekshirish
        previous_attempts = TestAttempt.objects.filter(
            student=student,
            test=test
        ).count()
        
        # Yangi urinish yaratish (qayta ishlash imkoniyati)
        new_attempt = TestAttempt.objects.create(
            student=student,
            test=test,
            attempt_number=previous_attempts + 1,
            is_retake=True
        )
        
        # Agar qayta ishlash so'rovi mavjud bo'lsa, uni tasdiqlangan deb belgilash
        retake_request = TestRetakeRequest.objects.filter(
            student=student,
            test=test,
            status='approved'
        ).first()
        
        if retake_request:
            retake_request.is_used = True
            retake_request.save()
        
        return JsonResponse({
            'message': f'{student.get_full_name()} uchun "{test.title}" testi qayta ochildi!',
            'attempt_id': new_attempt.id,
            'attempt_number': new_attempt.attempt_number
        })
        
    except Test.DoesNotExist:
        return JsonResponse({'error': 'Test topilmadi'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'O\'quvchi topilmadi'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def student_test_management(request):
    """Admin uchun o'quvchilarning test holatlarini boshqarish"""
    if request.user.role != 'admin':
        return redirect('accounts:dashboard')
    
    # Barcha faol testlar
    tests = Test.objects.filter(is_active=True)
    
    # Barcha tasdiqlangan o'quvchilar
    students = User.objects.filter(role='student', is_verified=True)
    
    # Har bir o'quvchi va test uchun urinishlar ma'lumotlari
    student_test_data = []
    
    for student in students:
        student_tests = []
        for test in tests:
            attempts = TestAttempt.objects.filter(student=student, test=test)
            latest_attempt = attempts.order_by('-started_at').first()
            
            # Qayta ishlash so'rovlari
            retake_requests = TestRetakeRequest.objects.filter(
                student=student,
                test=test
            ).order_by('-created_at')
            
            test_info = {
                'test': test,
                'attempts_count': attempts.count(),
                'latest_attempt': latest_attempt,
                'can_retake': latest_attempt is not None,
                'retake_requests': retake_requests
            }
            student_tests.append(test_info)
        
        student_test_data.append({
            'student': student,
            'tests': student_tests
        })
    
    context = {
        'student_test_data': student_test_data,
        'all_tests': tests
    }
    
    return render(request, 'tests_app/student_test_management.html', context)

@login_required
@require_http_methods(["POST", "DELETE"])
def delete_test_view(request, test_id):
    """Delete a test - Teachers (only their own tests) and Admins (any test)"""
    if request.user.role not in ['teacher', 'admin']:
        return JsonResponse({'error': 'Ruxsat berilmagan'}, status=403)
    
    # Teachers can only delete their own tests, admins can delete any test
    if request.user.role == 'teacher':
        test = get_object_or_404(Test, id=test_id, created_by=request.user)
    else:  # admin
        test = get_object_or_404(Test, id=test_id)
    
    try:
        test_title = test.title
        test.delete()
        return JsonResponse({
            'success': True,
            'message': f'"{test_title}" testi muvaffaqiyatli o\'chirildi!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Testni o\'chirishda xatolik: {str(e)}'
        }, status=500)

@login_required
def edit_test_view(request, test_id):
    """Edit an existing test and its questions (teachers and admins)"""
    # Teachers can only edit their own tests, admins can edit any test
    if request.user.role == 'teacher':
        test = get_object_or_404(Test, id=test_id, created_by=request.user)
    elif request.user.role == 'admin':
        test = get_object_or_404(Test, id=test_id)
    else:
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            # Check if it's FormData (for image upload) or JSON
            if request.content_type and 'multipart/form-data' in request.content_type:
                # Handle FormData for image upload or removal
                
                # Check if this is an image update/removal for existing question
                if request.POST.get('update_image_only') or request.POST.get('remove_image'):
                    question_id = request.POST.get('question_id')
                    if not question_id:
                        return JsonResponse({'success': False, 'error': 'Question ID kerak!'})
                    
                    try:
                        question = Question.objects.get(id=question_id, test=test)
                        
                        if request.POST.get('remove_image'):
                            # Remove image
                            if question.image:
                                question.image.delete(save=False)
                                question.image = None
                                question.save()
                                return JsonResponse({'success': True, 'message': 'Rasm o\'chirildi!'})
                            else:
                                return JsonResponse({'success': False, 'error': 'Rasm mavjud emas!'})
                        
                        elif request.POST.get('update_image_only'):
                            # Update image only
                            question_image = request.FILES.get('question_image')
                            if question_image:
                                # Remove old image if exists
                                if question.image:
                                    question.image.delete(save=False)
                                question.image = question_image
                                question.save()
                                return JsonResponse({
                                    'success': True, 
                                    'message': 'Rasm yuklandi!',
                                    'image_url': question.image.url
                                })
                            else:
                                return JsonResponse({'success': False, 'error': 'Rasm fayli topilmadi!'})
                    
                    except Question.DoesNotExist:
                        return JsonResponse({'success': False, 'error': 'Savol topilmadi!'})
                    except Exception as e:
                        return JsonResponse({'success': False, 'error': str(e)})
                
                # Handle new question creation with image
                question_text = request.POST.get('question_text')
                question_type = request.POST.get('question_type')
                points = request.POST.get('points', 1)
                explanation = request.POST.get('explanation', '')
                question_image = request.FILES.get('question_image')
                
                # Validation
                if not question_text:
                    return JsonResponse({
                        'success': False,
                        'error': 'Savol matni kiritilishi shart!'
                    })
                
                if not question_type:
                    return JsonResponse({
                        'success': False,
                        'error': 'Savol turi tanlanishi shart!'
                    })
                
                # Create new question with image
                try:
                    question = Question.objects.create(
                        test=test,
                        question_text=question_text,
                        question_type=question_type,
                        points=float(points),
                        order=test.questions.count() + 1,
                        explanation=explanation,
                        image=question_image
                    )
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': f'Question yaratishda xatolik: {str(e)}'
                    })
                
                # Handle choices
                try:
                    if question_type in ['single_choice', 'multiple_choice']:
                        choice_index = 0
                        while f'choices[{choice_index}][text]' in request.POST:
                            choice_text = request.POST.get(f'choices[{choice_index}][text]')
                            is_correct = request.POST.get(f'choices[{choice_index}][is_correct]') == 'true'
                            if choice_text:
                                Choice.objects.create(
                                    question=question,
                                    choice_text=choice_text,
                                    is_correct=is_correct
                                )
                            choice_index += 1
                except Exception:
                    # Don't return error here, question is already created
                    pass
                
                return JsonResponse({
                    'success': True,
                    'message': 'Savol muvaffaqiyatli qo\'shildi!',
                    'question_id': question.id
                })
            else:
                # Handle JSON data (existing functionality)
                data = json.loads(request.body)
            
                # Update test fields
                test.title = data.get('title', test.title)
                test.description = data.get('description', test.description)
                test.subject = data.get('subject', test.subject)
                test.grade = int(data.get('grade', test.grade))
                test.time_limit = int(data.get('time_limit', test.time_limit))
                test.max_attempts = int(data.get('max_attempts', test.max_attempts))
                test.show_results = data.get('show_results', test.show_results)
                test.is_active = data.get('is_active', test.is_active)
                test.shuffle_questions = data.get('shuffle_questions', test.shuffle_questions)
                test.save()

                # Update questions
                questions_data = data.get('questions', [])
            
                # Get existing question IDs
                existing_question_ids = set(test.questions.values_list('id', flat=True))
                new_question_ids = set()
            
                for i, q_data in enumerate(questions_data):
                    question_id = q_data.get('id')
                
                    if question_id and question_id in existing_question_ids:
                        # Update existing question
                        try:
                            question = Question.objects.get(id=question_id, test=test)
                            question.question_text = q_data['question_text']
                            question.question_type = q_data['question_type']
                            question.points = float(q_data.get('points', 1.0))
                            question.order = i + 1
                            question.explanation = q_data.get('explanation', '')
                            question.save()
                            new_question_ids.add(question_id)
                        
                            # Update choices
                            if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                                choices_data = q_data.get('choices', [])
                                # Clear existing choices
                                question.choices.all().delete()
                                # Add new choices
                                for c_data in choices_data:
                                    if c_data.get('text'):  # Only add if text is not empty
                                        Choice.objects.create(
                                            question=question,
                                            choice_text=c_data['text'],
                                            is_correct=c_data.get('is_correct', False)
                                        )
                            else:
                                # For text answers, remove all choices
                                question.choices.all().delete()
                            
                        except Question.DoesNotExist:
                            # If question doesn't exist, create new one
                            question = Question.objects.create(
                                test=test,
                                question_text=q_data['question_text'],
                                question_type=q_data['question_type'],
                                points=float(q_data.get('points', 1.0)),
                                order=i + 1,
                                explanation=q_data.get('explanation', '')
                            )
                            new_question_ids.add(question.id)
                        
                            if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                                for c_data in q_data.get('choices', []):
                                    if c_data.get('text'):
                                        Choice.objects.create(
                                            question=question,
                                            choice_text=c_data['text'],
                                            is_correct=c_data.get('is_correct', False)
                                        )
                    else:
                        # Create new question
                        question = Question.objects.create(
                            test=test,
                            question_text=q_data['question_text'],
                            question_type=q_data['question_type'],
                            points=float(q_data.get('points', 1.0)),
                            order=i + 1,
                            explanation=q_data.get('explanation', '')
                        )
                        new_question_ids.add(question.id)
                    
                        if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                            for c_data in q_data.get('choices', []):
                                if c_data.get('text'):
                                    Choice.objects.create(
                                        question=question,
                                        choice_text=c_data['text'],
                                        is_correct=c_data.get('is_correct', False)
                                    )
            
                # Remove questions that are no longer in the list
                questions_to_delete = existing_question_ids - new_question_ids
                if questions_to_delete:
                    test.questions.filter(id__in=questions_to_delete).delete()
                
                # Saqlangan savollarni JSON ko'rinishda qaytarish:
                questions = test.questions.all().order_by('order')
                questions_data = []
                for q in questions:
                    q_data = {
                        "id": q.id,
                        "question_text": q.question_text,
                        "question_type": q.question_type,
                        "points": q.points,
                        "explanation": q.explanation,
                        "image": q.image.url if q.image else None,
                        "choices": [
                            {"text": c.choice_text, "is_correct": c.is_correct}
                            for c in q.choices.all()
                        ]
                    }
                    questions_data.append(q_data)
                return JsonResponse({"success": True, "questions": questions_data})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET: Render edit page with test and questions
    if request.headers.get('Accept') == 'application/json':
        # Return JSON data for AJAX requests
        questions = test.questions.all().order_by('order')
        questions_data = []
        for q in questions:
            q_data = {
                'id': q.id,
                'question_text': q.question_text,
                'question_type': q.question_type,
                'points': q.points,
                'explanation': q.explanation,
                'image': q.image.url if q.image else None,  # Add image URL
                'choices': []
            }
            if q.question_type in ['single_choice', 'multiple_choice']:
                q_data['choices'] = [
                    {'id': c.id, 'text': c.choice_text, 'is_correct': c.is_correct}
                    for c in q.choices.all()
                ]
            questions_data.append(q_data)
        
        return JsonResponse({
            'test': {
                'id': test.id,
                'title': test.title,
                'subject': test.subject,
                'grade': test.grade,
                'time_limit': test.time_limit,
                'max_attempts': test.max_attempts,
                'description': test.description,
                'show_results': test.show_results,
                'is_active': test.is_active
            },
            'questions': questions_data
        })
    
    # Regular HTML request
    questions = test.questions.all().order_by('order')
    questions_data = []
    for q in questions:
        q_data = {
            'id': q.id,
            'question_text': q.question_text,
            'question_type': q.question_type,
            'points': q.points,
            'explanation': q.explanation,
            'image': q.image.url if q.image else None,  # Add image URL
            'choices': []
        }
        if q.question_type in ['single_choice', 'multiple_choice']:
            q_data['choices'] = [
                {'id': c.id, 'text': c.choice_text, 'is_correct': c.is_correct}
                for c in q.choices.all()
            ]
        questions_data.append(q_data)
    context = {
        'test': test,
        'questions': questions_data
    }
    return render(request, 'tests_app/edit_test.html', context)

@login_required
def start_test_view(request, test_id):
    """Admin tomonidan o'quvchi uchun testi boshlash"""
    test = get_object_or_404(Test, pk=test_id)
    questions = list(test.questions.all())
    random.shuffle(questions)  # Har bir o‘quvchi uchun random tartib
    
    context = {
        'test': test,
        'questions': questions,
    }
    return render(request, 'tests_app/start_test.html', context)

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .models import Test

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_teacher_tests(request):
    """
    Admin uchun: barcha testlarni ko'rsatish
    URL: /tests/admin/teacher-tests/
    """
    tests = (
        Test.objects
        .all()  # Barcha testlar
        .select_related('created_by')
        .annotate(attempts_count=Count('attempts', filter=Q(attempts__is_completed=True)))
        .order_by('-created_at')
    )

    return render(request, 'admin/tests_teacher_list.html', {
        'teacher_tests': tests
    })

@login_required
def export_subject_results_view(request):
    """Fan bo'yicha natijalarni Excel faylga export qilish"""
    if request.user.role not in ['teacher', 'admin']:
        return redirect('accounts:dashboard')
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'Excel export funksiyasi mavjud emas. Iltimos openpyxl kutubxonasini o\'rnating.'}, status=500)
    
    try:
        # Excel workbook yaratish
        wb = Workbook()
        
        # Barcha sinflarni olish (1-11)
        grades = list(range(1, 12))
        
        for grade in grades:
            # Har bir sinf uchun alohida sheet yaratish
            ws = wb.create_sheet(title=f"{grade}-sinf")
        
        # Header qo'shish
        ws['A1'] = f"{grade}-sinf Test Natijalari"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Umumiy ma'lumotlar
        ws['A3'] = "Umumiy Ma'lumotlar:"
        ws['A3'].font = Font(bold=True)
        
        # Bu sinf uchun testlar
        tests = Test.objects.filter(grade=grade, is_active=True)
        
        # Teacher faqat o'zi yaratgan testlarni ko'radi
        if request.user.role == 'teacher':
            tests = tests.filter(created_by=request.user)
        
        # Bu sinf o'quvchilari
        students = User.objects.filter(role='student', grade=grade, is_verified=True)
        
        # Bu sinf uchun test urinishlari - faqat har bir o'quvchining oxirgi natijasi
        from django.db.models import Max
        
        # Har bir o'quvchi uchun eng so'nggi attempt ID'sini topish
        latest_attempts_filter = {
            'test__grade': grade,
            'student__grade': grade,
            'is_completed': True
        }
        
        # Teacher faqat o'zi yaratgan testlar natijalarini ko'radi
        if request.user.role == 'teacher':
            latest_attempts_filter['test__created_by'] = request.user
        
        latest_attempts = TestAttempt.objects.filter(
            **latest_attempts_filter
        ).values('student').annotate(
            latest_attempt_id=Max('id')
        ).values_list('latest_attempt_id', flat=True)
        
        # Faqat oxirgi attempt'larni olish
        attempts = TestAttempt.objects.filter(
            id__in=latest_attempts
        ).select_related('student', 'test', 'result')
        
        # Statistika hisoblash
        total_students = students.count()
        total_tests = tests.count()
        total_attempts = attempts.count()
        
        if total_attempts > 0:
            avg_percentage = attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
            highest_score = attempts.aggregate(max=Max('percentage'))['max'] or 0
            lowest_score = attempts.aggregate(min=Min('percentage'))['min'] or 0
        else:
            avg_percentage = 0
            highest_score = 0
            lowest_score = 0
        
        # Ma'lumotlarni yozish
        ws['A4'] = f"Jami O'quvchilar: {total_students}"
        ws['A5'] = f"Jami Testlar: {total_tests}"
        ws['A6'] = f"Jami Urinishlar: {total_attempts}"
        ws['A7'] = f"O'rtacha Ball: {avg_percentage:.1f}%"
        ws['A8'] = f"Eng Yuqori Ball: {highest_score:.1f}%"
        ws['A9'] = f"Eng Past Ball: {lowest_score:.1f}%"
        
        # Bo'sh qator
        ws['A11'] = "Barcha Natijalar:"
        ws['A11'].font = Font(bold=True)
        
        # Header qator
        headers = ['O\'quvchi', 'Test', 'Ball', 'Foiz', 'Vaqt', 'Sana']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Natijalarni yozish
        row = 13
        for attempt in attempts.order_by('-percentage'):
            student_name = f"{attempt.student.first_name} {attempt.student.last_name}"
            if not student_name.strip():
                student_name = attempt.student.username
            
            ws.cell(row=row, column=1, value=student_name)
            ws.cell(row=row, column=2, value=attempt.test.title)
            ws.cell(row=row, column=3, value=attempt.score)
            ws.cell(row=row, column=4, value=f"{attempt.percentage:.1f}%")
            ws.cell(row=row, column=5, value=str(attempt.time_taken))
            ws.cell(row=row, column=6, value=attempt.finished_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Ball bo'yicha rang berish
            if attempt.percentage >= 81:
                fill_color = "C6EFCE"  # Yashil
            elif attempt.percentage >= 51:
                fill_color = "FFEB9C"  # Sariq
            elif attempt.percentage >= 31:
                fill_color = "FFC7CE"  # Qizil
            else:
                fill_color = "FFC7CE"  # Qizil
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row += 1
        
        # Ustun kengliklarini sozlash
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
    
        # Default sheet'ni o'chirish
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sinf_natijalari.xlsx"'
        
        # Excel faylni saqlash
        wb.save(response)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Export xatolik yuz berdi: {str(e)}'}, status=500)

@login_required
def students_test_results_view(request):
    """O'quvchilar yechgan testlarning natijalarini ko'rsatish - Admin va Teacher uchun"""
    from django.db.models import Q
    from tests_app.models import TestAttempt, Test
    
    if request.user.role not in ['admin', 'teacher']:
        return redirect('accounts:dashboard')
    
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        try:
            # Filter parametrlari
            grade_filter = request.GET.get('grade', '')
            subject_filter = request.GET.get('subject', '')
            test_id_filter = request.GET.get('test', '')
            student_filter = request.GET.get('student', '')
            
            # ORM query boshlandi
            attempts = TestAttempt.objects.filter(
        is_completed=True
            ).select_related('student', 'test', 'test__created_by')
            
            # Teacher uchun faqat o'z yaratgan testlarini ko'rsatish
            if request.user.role == 'teacher':
                attempts = attempts.filter(test__created_by=request.user)
            
            # Filterlar
            if grade_filter:
                attempts = attempts.filter(student__grade=grade_filter)
            if subject_filter:
                attempts = attempts.filter(test__subject=subject_filter)
            if test_id_filter:
                attempts = attempts.filter(test_id=test_id_filter)
            if student_filter:
                attempts = attempts.filter(
                    Q(student__first_name__icontains=student_filter) |
                    Q(student__last_name__icontains=student_filter) |
                    Q(student__username__icontains=student_filter)
                )
            
            # Tartib - eng yaxshi natijani topish uchun avval percentage bo'yicha sort qilamiz
            attempts = attempts.order_by('student__id', 'test__id', '-percentage', '-finished_at')
            
            # Har bir o'quvchi-test juftligi uchun faqat eng yaxshi natijani olish
            seen_combinations = set()
            results_data = []
            for attempt in attempts:
                # O'quvchi-test juftligi kaliti
                combination_key = (attempt.student.id, attempt.test.id)
                
                # Agar bu juftlik ko'rilgan bo'lsa, o'tkazib yuborish
                if combination_key in seen_combinations:
                    continue
                
                # Juftlikni qo'shish
                seen_combinations.add(combination_key)
                percentage = attempt.percentage or 0
                
                # Baholash
                if percentage >= 81:
                    grade_text = "A'lo"
                elif percentage >= 51:
                    grade_text = "Yaxshi"
                elif percentage >= 31:
                    grade_text = "Qoniqarli"
                else:
                    grade_text = "Qoniqarsiz"
                
                results_data.append({
                    'id': attempt.id,
                    'test': {
                        'id': attempt.test.id,
                        'title': attempt.test.title,
                        'subject': attempt.test.subject,
                        'grade': attempt.test.grade
                    },
                    'student': {
                        'id': attempt.student.id,
                        'username': attempt.student.username,
                        'first_name': attempt.student.first_name or '',
                        'last_name': attempt.student.last_name or '',
                        'student_id': getattr(attempt.student, 'student_id', ''),
                        'class_name': getattr(attempt.student, 'class_name', ''),
                        'grade': attempt.student.grade or '',
                        'full_name': attempt.student.get_full_name() or attempt.student.username
                    },
                    'score': attempt.score or 0,
                    'total_points': attempt.total_points or 0,
                    'percentage': percentage,
                    'grade': grade_text,
                    'time_taken': str(attempt.time_taken) if attempt.time_taken else '',
                    'finished_at': attempt.finished_at.isoformat() if attempt.finished_at else None,
                    'started_at': attempt.started_at.isoformat() if attempt.started_at else None,
                    'attempt_number': attempt.attempt_number,
                })
            
            # Natijalarni to'g'ri tartiblash (grade, last_name, first_name bo'yicha)
            results_data.sort(key=lambda x: (
                x['student']['grade'] or '',
                x['student']['last_name'] or '',
                x['student']['first_name'] or '',
                x['test']['subject'] or '',
                x['test']['title'] or ''
            ))
            
            # Statistika
            stats = {
                'total_results': len(results_data),
                'avg_percentage': sum([r['percentage'] for r in results_data]) / len(results_data) if results_data else 0,
                'excellent_count': sum(1 for r in results_data if r['percentage'] >= 81),
                'good_count': sum(1 for r in results_data if 61 <= r['percentage'] < 81),
                'satisfactory_count': sum(1 for r in results_data if 31 <= r['percentage'] < 61),
                'poor_count': sum(1 for r in results_data if r['percentage'] < 31),
            }
            
            return JsonResponse({
                'results': results_data,
                'stats': stats
            }, json_dumps_params={'ensure_ascii': False})
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'error': f'Xatolik: {str(e)}',
                'results': [],
                'stats': {
                    'total_results': 0,
                    'avg_percentage': 0,
                    'excellent_count': 0,
                    'poor_count': 0
                }
            }, status=500, json_dumps_params={'ensure_ascii': False})
    
    # HTML render - O'qituvchi uchun faqat o'zining fanlarini ko'rsatish
    teacher_subjects = []
    if request.user.role == 'teacher':
        # O'qituvchining fani (ro'yxatdan o'tgan fani)
        if request.user.subject:
            teacher_subjects.append(request.user.subject)
        
        # O'qituvchi yaratgan testlarning barcha fanlari
        created_test_subjects = Test.objects.filter(
            created_by=request.user
        ).values_list('subject', flat=True).distinct()
        
        # Barcha fanlarni birlashtirish (takrorlanishlarsiz)
        all_subjects = set(teacher_subjects) | set(created_test_subjects)
        teacher_subjects = sorted(list(all_subjects))
    
    return render(request, 'tests_app/students_test_results.html', {
        'user_role': request.user.role,
        'teacher_subjects': teacher_subjects if request.user.role == 'teacher' else []
    })

@login_required
def export_students_test_results_view(request):
    """O'quvchilar test natijalarini Excel formatida export qilish - Sinflar aro - Admin va Teacher uchun"""
    if request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)
    
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'Excel export funksiyasi mavjud emas. Iltimos openpyxl kutubxonasini o\'rnating.'}, status=500)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Filter parametrlari
        grade_filter = request.GET.get('grade', '')
        subject_filter = request.GET.get('subject', '')
        test_id_filter = request.GET.get('test', '')
        
        # Excel workbook yaratish
        wb = Workbook()
        
        # Barcha sinflar uchun (1-11)
        grades = list(range(1, 12))
        
        # Har bir sinf uchun alohida sheet yaratish
        for grade in grades:
            if grade_filter and int(grade_filter) != grade:
                continue
                
            # SQL query - parametrlarni to'g'ri ishlatish
            # Django ORM ishlatish - xavfsiz va to'g'ri
            attempts_query = TestAttempt.objects.filter(
                is_completed=True,
                student__grade=grade
            ).select_related('student', 'test')
            
            # Teacher uchun faqat o'z yaratgan testlarini ko'rsatish
            if request.user.role == 'teacher':
                attempts_query = attempts_query.filter(test__created_by=request.user)
            
            # Filterlar
            if subject_filter:
                attempts_query = attempts_query.filter(test__subject=subject_filter)
            if test_id_filter:
                attempts_query = attempts_query.filter(test_id=int(test_id_filter))
            
            # Har bir o'quvchi-test juftligi uchun faqat eng yaxshi natijani olish
            # Avval percentage bo'yicha sort qilamiz (eng yaxshi birinchi)
            attempts_query = attempts_query.order_by('student__id', 'test__id', '-percentage', '-finished_at')
            
            attempts_data = []
            seen_combinations = set()
            for attempt in attempts_query:
                # O'quvchi-test juftligi kaliti
                combination_key = (attempt.student.id, attempt.test.id)
                
                # Agar bu juftlik ko'rilgan bo'lsa, o'tkazib yuborish (faqat birinchi eng yaxshi natija)
                if combination_key in seen_combinations:
                    continue
                
                # Juftlikni qo'shish
                seen_combinations.add(combination_key)
                attempts_data.append({
                    'id': attempt.id,
                    'percentage': attempt.percentage,
                    'score': attempt.score,
                    'total_points': attempt.total_points,
                    'finished_at': attempt.finished_at,
                    'started_at': attempt.started_at,
                    'attempt_number': attempt.attempt_number,
                    'time_taken': attempt.time_taken,
                    'first_name': attempt.student.first_name if attempt.student else '',
                    'last_name': attempt.student.last_name if attempt.student else '',
                    'username': attempt.student.username if attempt.student else '',
                    'student_unique_id': attempt.student.student_id if attempt.student else '',
                    'class_name': attempt.student.class_name if attempt.student else '',
                    'test_title': attempt.test.title if attempt.test else '',
                    'subject': attempt.test.subject if attempt.test else ''
                })
            
            # Natijalarni to'g'ri tartiblash (test, student bo'yicha)
            attempts_data.sort(key=lambda x: (
                x.get('subject', '') or '',
                x.get('test_title', '') or '',
                x.get('last_name', '') or '',
                x.get('first_name', '') or ''
            ))
            
            # Agar bu sinf uchun natijalar bo'lmasa, sheet yaratma
            if not attempts_data:
                continue
            
            # Sheet yaratish
            ws = wb.create_sheet(title=f"{grade}-sinf")
            
            # Header - Chiroyli dizayn
            ws['A1'] = f"📊 {grade}-sinf O'quvchilar Test Natijalari"
            ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="7C4DFF", end_color="9C27B0", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A1:K1')
            ws.row_dimensions[1].height = 35
            
            # Statistika - Chiroyli dizayn
            stats_row = 3
            ws[f'A{stats_row}'] = "📈 Statistika"
            ws[f'A{stats_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{stats_row}'].fill = PatternFill(start_color="00E676", end_color="00C853", fill_type="solid")
            ws[f'A{stats_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'A{stats_row}:K{stats_row}')
            ws.row_dimensions[stats_row].height = 30
            
            if attempts_data:
                avg_percentage = sum([r['percentage'] or 0 for r in attempts_data]) / len(attempts_data)
                excellent_count = sum(1 for r in attempts_data if (r['percentage'] or 0) >= 81)
                good_count = sum(1 for r in attempts_data if 61 <= (r['percentage'] or 0) < 81)
                satisfactory_count = sum(1 for r in attempts_data if 31 <= (r['percentage'] or 0) < 61)
                poor_count = sum(1 for r in attempts_data if (r['percentage'] or 0) < 31)
                
                # Statistika - Soddalashtirilgan ko'rinish
                stats_row_start = stats_row + 1
                ws[f'A{stats_row_start}'] = f"📊 Jami Natijalar: {len(attempts_data)}"
                ws[f'D{stats_row_start}'] = f"📈 O'rtacha Foiz: {avg_percentage:.1f}%"
                ws[f'G{stats_row_start}'] = f"⭐ A'lo: {excellent_count} | ✅ Yaxshi: {good_count}"
                
                ws[f'A{stats_row_start + 1}'] = f"⚡ Qoniqarli: {satisfactory_count} | ❌ Qoniqarsiz: {poor_count}"
                
                # Statistikalar uchun format
                for r in range(stats_row_start, stats_row_start + 2):
                    for col in range(1, 12):
                        cell = ws.cell(row=r, column=col)
                        if r == stats_row_start:
                            if col == 1:
                                cell.font = Font(bold=True, size=11, color="667EEA")
                            elif col == 4:
                                cell.font = Font(bold=True, size=11, color="00E676")
                            elif col == 7:
                                cell.font = Font(bold=True, size=11, color="4CAF50")
                            else:
                                cell.font = Font(bold=True, size=11)
                        else:
                            if col == 1:
                                cell.font = Font(bold=True, size=11, color="FF9800")
                            else:
                                cell.font = Font(bold=True, size=11, color="F44336")
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                ws.row_dimensions[stats_row_start].height = 25
                ws.row_dimensions[stats_row_start + 1].height = 25
            
            header_row = stats_row + 4
            
            # Jadval header - Chiroyli dizayn
            headers = ['№', 'O\'quvchi', 'Student ID', 'Sinf', 'Test', 'Fan', 'Ball', 'Foiz', 'Baholash', 'Urinish', 'Sana']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='medium', color="FFFFFF"),
                    right=Side(style='medium', color="FFFFFF"),
                    top=Side(style='medium', color="FFFFFF"),
                    bottom=Side(style='medium', color="FFFFFF")
                )
            ws.row_dimensions[header_row].height = 30
            
            # Ma'lumotlarni yozish
            row = header_row + 1
            for idx, attempt in enumerate(attempts_data, 1):
                percentage = attempt['percentage'] or 0

                # Baholash - Yorqin ranglar
                if percentage >= 81:
                    grade_text = "A'lo"
                    fill_color = "00FF88"  # Yorqin yashil
                    text_color = "FFFFFF"
                elif percentage >= 51:
                    grade_text = "Yaxshi"
                    fill_color = "FFD700"  # Yorqin sariq
                    text_color = "000000"
                elif percentage >= 31:
                    grade_text = "Qoniqarli"
                    fill_color = "FF9800"  # Yorqin to'q sariq
                    text_color = "FFFFFF"
                else:
                    grade_text = "Qoniqarsiz"
                    fill_color = "FF5252"  # Yorqin qizil
                    text_color = "FFFFFF"
                
                student_name = f"{attempt['first_name'] or ''} {attempt['last_name'] or ''}".strip()
                if not student_name:
                    student_name = attempt['username']
                
                finished_date = attempt['finished_at'].strftime('%Y-%m-%d %H:%M:%S') if attempt['finished_at'] else ''
                
                data = [
                    idx,
                    student_name,
                    attempt['student_unique_id'] or '',
                    attempt['class_name'] or '',
                    attempt['test_title'] or '',
                    attempt['subject'] or '',
                    attempt['score'] or 0,
                    percentage,
                    grade_text,
                    attempt['attempt_number'] or 1,
                    finished_date
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Har bir qatorning rangini alternativ qilish (zebra striping)
                    if idx % 2 == 0:
                        row_fill_color = "F5F5F5"  # Oq-grey
                    else:
                        row_fill_color = "FFFFFF"  # Oq
                    
                    # Baholash ustuniga rang berish
                    if col == 9:  # Baholash ustuni
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.font = Font(bold=True, size=11, color=text_color)
                    else:
                        cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")
                    
                    # Alignment
                    if col in [1, 7, 8, 10]:  # Raqamli ustunlar
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Border - Chiroyli
                    cell.border = Border(
                        left=Side(style='thin', color="E0E0E0"),
                        right=Side(style='thin', color="E0E0E0"),
                        top=Side(style='thin', color="E0E0E0"),
                        bottom=Side(style='thin', color="E0E0E0")
                    )
                
                # Foiz ustuniga rang berish
                percentage_cell = ws.cell(row=row, column=8)
                if percentage >= 81:
                    percentage_cell.font = Font(bold=True, size=11, color="00C853")
                elif percentage >= 51:
                    percentage_cell.font = Font(bold=True, size=11, color="FF9800")
                elif percentage >= 31:
                    percentage_cell.font = Font(bold=True, size=11, color="FF6F00")
                else:
                    percentage_cell.font = Font(bold=True, size=11, color="F44336")
                
                row += 1
            
            # Ustun kengliklarini sozlash - Optimal kengliklar
            column_widths = [6, 28, 15, 12, 35, 18, 12, 12, 15, 10, 22]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Jadval uchun freeze panes (header har doim ko'rinib turadi)
            last_data_row = row - 1  # Oxirgi ma'lumot qatori
            try:
                if last_data_row > header_row:
                    freeze_cell = f'A{header_row + 1}'
                    ws.freeze_panes = freeze_cell
            except Exception:
                pass
            
            # Auto filter qo'shish
            try:
                if last_data_row > header_row:
                    last_col_letter = 'K'  # 11-ustun
                    auto_filter_range = f'A{header_row}:{last_col_letter}{last_data_row}'
                    ws.auto_filter.ref = auto_filter_range
            except Exception:
                pass
        
        # Default sheet'ni o'chirish
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Filename yaratish
        if grade_filter:
            # Agar faqat bitta sinf tanlangan bo'lsa, o'sha sinf nomi bilan fayl yaratamiz
            filename = f"{grade_filter}-sinf_test_natijalari.xlsx"
        else:
            # Agar barcha sinflar tanlangan bo'lsa, umumiy nom bilan
            filename = "barcha_sinflar_test_natijalari.xlsx"
        
        # Response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Excel faylni saqlash
        wb.save(response)
        
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Export xatolik yuz berdi: {str(e)}'}, status=500)

# export_all_results_view funksiyasi olib tashlandi - /tests/export-all-results/ bo'limi kerak emas edi

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def export_subject_results_view(request):
    """Fanlar bo'yicha natijalarni Excel faylga export qilish - har bir fan alohida sheet"""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'Excel export funksiyasi mavjud emas.'}, status=500)
    
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
        
        # Excel workbook yaratish
        wb = Workbook()
        wb.remove(wb.active)  # Default sheet'ni o'chirish
        
        # Barcha fanlarni olish
        subjects_query = Test.objects.all()
        
        # Teacher faqat o'zi yaratgan testlarni ko'radi
        if request.user.role == 'teacher':
            subjects_query = subjects_query.filter(created_by=request.user)
        
        subjects = subjects_query.values_list('subject', flat=True).distinct().order_by('subject')
        
        for subject in subjects:
            if not subject:
                continue
                
            
            # Sheet yaratish
            ws = wb.create_sheet(title=subject[:31])  # Excel sheet name limit: 31 characters
            
            # Header qo'shish
            ws['A1'] = f"{subject} - Barcha Natijalar"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Umumiy statistika
            tests = Test.objects.filter(subject=subject, is_active=True)
            
            # Teacher faqat o'zi yaratgan testlarni ko'radi
            if request.user.role == 'teacher':
                tests = tests.filter(created_by=request.user)
            
            total_tests = tests.count()
            
            attempts = TestAttempt.objects.filter(
                test__subject=subject,
                is_completed=True
            ).select_related('student', 'test')
            
            # Teacher faqat o'zi yaratgan testlar natijalarini ko'radi
            if request.user.role == 'teacher':
                attempts = attempts.filter(test__created_by=request.user)
            
            total_attempts = attempts.count()
            
            if total_attempts > 0:
                avg_percentage = attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
                max_percentage = attempts.aggregate(max=Max('percentage'))['max'] or 0
                min_percentage = attempts.aggregate(min=Min('percentage'))['min'] or 0
            else:
                avg_percentage = max_percentage = min_percentage = 0
            
            # Statistika qo'shish
            ws['A3'] = "Umumiy Ma'lumotlar:"
            ws['A3'].font = Font(bold=True, size=12)
            
            ws['A4'] = f"Jami Testlar: {total_tests}"
            ws['A5'] = f"Jami Urinishlar: {total_attempts}"
            ws['A6'] = f"O'rtacha Foiz: {avg_percentage:.1f}%"
            ws['A7'] = f"Eng Yuqori Foiz: {max_percentage:.1f}%"
            ws['A8'] = f"Eng Past Foiz: {min_percentage:.1f}%"
            
            # Natijalar jadvali header
            ws['A10'] = "Barcha Natijalar:"
            ws['A10'].font = Font(bold=True, size=12)
            
            headers = ['O\'quvchi', 'Sinf', 'Test Nomi', 'Ball', 'Maksimal', 'Foiz', 'Vaqt', 'Sana']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=11, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Natijalarni yozish
            row = 12
            for attempt in attempts.order_by('-percentage', '-finished_at'):
                student_name = f"{attempt.student.first_name} {attempt.student.last_name}".strip()
                if not student_name:
                    student_name = attempt.student.username
                
                ws.cell(row=row, column=1, value=student_name)
                ws.cell(row=row, column=2, value=attempt.student.grade or '')
                ws.cell(row=row, column=3, value=attempt.test.title)
                ws.cell(row=row, column=4, value=attempt.score)
                ws.cell(row=row, column=5, value=attempt.total_points)
                ws.cell(row=row, column=6, value=f"{attempt.percentage:.1f}%")
                ws.cell(row=row, column=7, value=str(attempt.time_taken) if attempt.time_taken else '-')
                ws.cell(row=row, column=8, value=attempt.finished_at.strftime('%d.%m.%Y %H:%M') if attempt.finished_at else '-')
                
                # Rangli formatlar
                if attempt.percentage >= 81:
                    fill_color = "C6EFCE"  # Yashil
                elif attempt.percentage >= 51:
                    fill_color = "FFEB9C"  # Sariq
                elif attempt.percentage >= 31:
                    fill_color = "BDD7EE"  # Ko'k
                else:
                    fill_color = "FFC7CE"  # Qizil
                
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
            
            # Ustun kengliklarini sozlash
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 18
        
        # Agar hech qanday fan topilmasa
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="Ma'lumot yo'q")
            ws['A1'] = "Hozircha natijalar mavjud emas"
            ws['A1'].font = Font(bold=True, size=14)
        
        # Response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="fanlar_boyicha_natijalar.xlsx"'
        
        # Excel faylni saqlash
        wb.save(response)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'Export xatolik yuz berdi: {str(e)}'}, status=500)

@login_required
@user_passes_test(lambda u: u.role == 'admin')
@require_http_methods(["POST"])
def delete_all_tests_view(request):
    """Barcha testlarni o'chirish (faqat admin) - O'quvchilar saqlanadi"""
    try:
        # Hisob-kitob
        total_tests = Test.objects.count()
        total_questions = Question.objects.count()
        total_choices = Choice.objects.count()
        total_attempts = TestAttempt.objects.count()
        total_answers = Answer.objects.count()
        total_results = TestResult.objects.count()
        total_retake_requests = TestRetakeRequest.objects.count()
        total_users = User.objects.count()
        
        # Barcha testlarni o'chirish (CASCADE bo'lgani uchun bog'liq obyektlar avtomatik o'chadi)
        deleted_count = Test.objects.all().delete()[0]
        
        # Tekshirish
        remaining_users = User.objects.count()
        
        return JsonResponse({
            'success': True,
            'message': f'Barcha testlar muvaffaqiyatli o\'chirildi!',
            'deleted': {
                'tests': total_tests,
                'questions': total_questions,
                'choices': total_choices,
                'attempts': total_attempts,
                'answers': total_answers,
                'results': total_results,
                'retake_requests': total_retake_requests,
                'total_deleted': deleted_count
            },
            'preserved': {
                'users': total_users,
                'users_after': remaining_users
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Testlarni o\'chirishda xatolik: {str(e)}'
        }, status=500)
