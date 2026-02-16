"""
Real-time monitoring views
O'quvchilar bo'yicha monitoring, test sessiyalari, vaqt boshqarish va Excel export
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db.models import Avg, Max, Min
import json
from .models import Test, TestAttempt, TestResult
from accounts.models import User

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def active_test_sessions_view(request, test_id):
    """Test uchun aktiv sessiyalar"""
    test = get_object_or_404(Test, id=test_id)
    
    if request.headers.get('Accept') == 'application/json':
        active_attempts = TestAttempt.objects.filter(
            test=test,
            is_completed=False,
            is_terminated=False
        ).select_related('student').order_by('-started_at')
        
        sessions_data = []
        for attempt in active_attempts:
            elapsed = timezone.now() - attempt.started_at
            elapsed_min = int(elapsed.total_seconds() / 60)
            remaining_min = max(0, test.time_limit - elapsed_min)
            answered = attempt.answers.count()
            progress = int((answered / test.total_questions * 100)) if test.total_questions > 0 else 0
            
            sessions_data.append({
                'attempt_id': attempt.id,
                'student_id': attempt.student.id,
                'student_name': attempt.student.get_full_name() or attempt.student.username,
                'student_username': attempt.student.username,
                'started_at': attempt.started_at.strftime('%H:%M:%S'),
                'elapsed_minutes': elapsed_min,
                'remaining_minutes': remaining_min,
                'answered_count': answered,
                'total_questions': test.total_questions,
                'progress_percent': progress,
                'current_question': attempt.current_question_index + 1,
                'is_late': elapsed_min > test.time_limit,
            })
        
        return JsonResponse({
            'sessions': sessions_data,
            'total_active': len(sessions_data),
            'test': {
                'id': test.id,
                'title': test.title,
                'time_limit': test.time_limit,
                'total_questions': test.total_questions
            }
        })
    
    return render(request, 'tests_app/active_sessions.html', {'test': test})

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
@require_http_methods(["POST"])
def terminate_test_attempt_view(request, attempt_id):
    """Testni to'xtatish"""
    attempt = get_object_or_404(TestAttempt, id=attempt_id)
    
    if request.user.role not in ['teacher', 'admin']:
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)
    
    try:
        data = json.loads(request.body)
        reason = data.get('reason', 'O\'qituvchi tomonidan to\'xtatildi')
        
        attempt.is_terminated = True
        attempt.terminated_by = request.user
        attempt.termination_reason = reason
        attempt.is_completed = True
        attempt.finished_at = timezone.now()
        attempt.time_taken = timezone.now() - attempt.started_at
        results = attempt.calculate_score()
        attempt.save()
        
        result, created = TestResult.objects.get_or_create(
            attempt=attempt,
            defaults={
                'correct_answers': results.get('correct_answers', 0),
                'incorrect_answers': results.get('incorrect_answers', 0),
                'unanswered': results.get('unanswered', 0),
                'grade': ''
            }
        )
        result.grade = result.calculate_grade()
        result.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Test to\'xtatildi',
            'attempt': {
                'id': attempt.id,
                'student_name': attempt.student.get_full_name() or attempt.student.username,
                'score': attempt.score,
                'percentage': round(attempt.percentage, 1),
                'terminated_at': attempt.finished_at.strftime('%H:%M:%S'),
                'reason': reason
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def student_test_detail_view(request, attempt_id):
    """O'quvchi test detali"""
    attempt = get_object_or_404(TestAttempt, id=attempt_id)
    test = attempt.test
    
    if request.headers.get('Accept') == 'application/json':
        questions_data = []
        for question in test.questions.all():
            answer = attempt.answers.filter(question=question).first()
            
            choices_data = []
            for choice in question.choices.all():
                choices_data.append({
                    'id': choice.id,
                    'text': choice.choice_text,
                    'is_correct': choice.is_correct,
                    'is_selected': answer and choice in answer.selected_choices.all() if answer else False
                })
            
            questions_data.append({
                'id': question.id,
                'order': question.order,
                'text': question.question_text,
                'points': question.points,
                'type': question.question_type,
                'choices': choices_data,
                'is_answered': answer is not None,
                'is_correct': answer.is_correct() if answer else False,
                'student_answer': answer.get_student_answer_text() if answer else 'Javob berilmagan',
                'correct_answer': question.get_correct_answer_text()
            })
        
        return JsonResponse({
            'attempt': {
                'id': attempt.id,
                'student_name': attempt.student.get_full_name() or attempt.student.username,
                'started_at': attempt.started_at.strftime('%d.%m.%Y %H:%M'),
                'is_completed': attempt.is_completed,
                'is_terminated': attempt.is_terminated,
                'current_question': attempt.current_question_index + 1,
                'answered_count': attempt.answers.count(),
                'total_questions': test.total_questions,
                'score': attempt.score,
                'percentage': round(attempt.percentage, 1) if attempt.percentage else 0
            },
            'test': {
                'id': test.id,
                'title': test.title,
                'subject': test.subject,
                'time_limit': test.time_limit
            },
            'questions': questions_data
        })
    
    return render(request, 'tests_app/student_test_detail.html', {
        'attempt': attempt,
        'test': test
    })

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
@require_http_methods(["POST"])
def control_time_view(request, attempt_id):
    """Vaqtni boshqarish"""
    attempt = get_object_or_404(TestAttempt, id=attempt_id)
    
    if request.user.role not in ['teacher', 'admin']:
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)
    
    if attempt.is_completed:
        return JsonResponse({'error': 'Test tugagan'}, status=400)
    
    try:
        data = json.loads(request.body)
        action = data.get('action')
        minutes = int(data.get('minutes', 0))
        
        test = attempt.test
        elapsed = timezone.now() - attempt.started_at
        elapsed_min = int(elapsed.total_seconds() / 60)
        current_remaining = test.time_limit - elapsed_min
        
        if action == 'add':
            test.time_limit += minutes
            test.save()
            new_remaining = current_remaining + minutes
            message = f'{minutes} daqiqa qo\'shildi'
            
        elif action == 'reduce':
            if minutes >= current_remaining:
                return JsonResponse({'error': 'Juda ko\'p'}, status=400)
            test.time_limit -= minutes
            test.save()
            new_remaining = current_remaining - minutes
            message = f'{minutes} daqiqa kamaytirildi'
            
        elif action == 'remove':
            test.time_limit = 999999
            test.save()
            new_remaining = 999999
            message = 'Vaqt cheklovi olib tashlandi'
            
        else:
            return JsonResponse({'error': 'Noto\'g\'ri amal'}, status=400)
        
        return JsonResponse({
            'success': True,
            'message': message,
            'new_time': new_remaining if new_remaining < 999999 else 'Cheksiz',
            'action': action,
            'minutes': minutes
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def students_monitoring_view(request):
    """O'quvchilar bo'yicha monitoring"""
    
    if request.headers.get('Accept') == 'application/json':
        students = User.objects.filter(role='student').order_by('grade', 'last_name')
        
        students_data = []
        total_active = 0
        total_completed = 0
        
        for student in students:
            active_attempts = TestAttempt.objects.filter(
                student=student,
                is_completed=False,
                is_terminated=False
            ).select_related('test').order_by('-started_at')
            
            completed_count = TestAttempt.objects.filter(
                student=student,
                is_completed=True
            ).count()
            
            active_tests_data = []
            for attempt in active_attempts:
                test = attempt.test
                elapsed = timezone.now() - attempt.started_at
                elapsed_min = int(elapsed.total_seconds() / 60)
                remaining_min = max(0, test.time_limit - elapsed_min)
                answered = attempt.answers.count()
                progress = int((answered / test.total_questions * 100)) if test.total_questions > 0 else 0
                
                active_tests_data.append({
                    'attempt_id': attempt.id,
                    'test_id': test.id,
                    'test_title': test.title,
                    'test_subject': test.subject,
                    'total_questions': test.total_questions,
                    'answered': answered,
                    'current_question': attempt.current_question_index + 1,
                    'elapsed_minutes': elapsed_min,
                    'remaining_minutes': remaining_min,
                    'progress_percent': progress,
                    'is_late': elapsed_min > test.time_limit,
                    'started_at': attempt.started_at.strftime('%H:%M')
                })
                total_active += 1
            
            total_completed += completed_count
            
            if active_tests_data or completed_count > 0:
                students_data.append({
                    'id': student.id,
                    'name': student.get_full_name() or student.username,
                    'username': student.username,
                    'grade': student.grade,
                    'active_tests': active_tests_data,
                    'completed_count': completed_count
                })
        
        return JsonResponse({
            'students': students_data,
            'stats': {
                'total_students': len(students_data),
                'active_sessions': total_active,
                'completed_tests': total_completed
            }
        })
    
    return render(request, 'tests_app/student_monitoring.html')

@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def export_cross_grade_results(request):
    """Sinflararo Excel export"""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({'error': 'openpyxl kerak'}, status=500)
    
    try:
        wb = Workbook()
        wb.remove(wb.active)
        grades = list(range(4, 10))
        
        # Summary sheet
        ws = wb.create_sheet(title="Umumiy")
        ws['A1'] = "SINFLARARO NATIJALAR"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        headers = ['Sinf', 'O\'quvchilar', 'Testlar', 'Urinishlar', 'O\'rtacha', 'Max', 'Min', 'O\'tish']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        row = 4
        for grade in grades:
            attempts = TestAttempt.objects.filter(
                test__grade=grade, student__grade=grade, is_completed=True
            )
            total = attempts.count()
            
            if total > 0:
                avg = attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
                mx = attempts.aggregate(max=Max('percentage'))['max'] or 0
                mn = attempts.aggregate(min=Min('percentage'))['min'] or 0
                pas = attempts.filter(percentage__gte=60).count() / total * 100
            else:
                avg = mx = mn = pas = 0
            
            ws.cell(row=row, column=1, value=f"{grade}-sinf")
            ws.cell(row=row, column=2, value=User.objects.filter(role='student', grade=grade).count())
            ws.cell(row=row, column=3, value=Test.objects.filter(grade=grade).count())
            ws.cell(row=row, column=4, value=total)
            ws.cell(row=row, column=5, value=f"{avg:.1f}%")
            ws.cell(row=row, column=6, value=f"{mx:.1f}%")
            ws.cell(row=row, column=7, value=f"{mn:.1f}%")
            ws.cell(row=row, column=8, value=f"{pas:.1f}%")
            
            fill = "C6EFCE" if avg >= 80 else "FFEB9C" if avg >= 60 else "FFC7CE"
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            row += 1
        
        # Detail sheets
        for grade in grades:
            s = wb.create_sheet(title=f"{grade}-sinf")
            s['A1'] = f"{grade}-SINF"
            s['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            s['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            s.merge_cells('A1:G1')
            
            for col, h in enumerate(['O\'quvchi', 'Fan', 'Test', 'Ball', 'Foiz', 'Holat', 'Sana'], 1):
                c = s.cell(row=3, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            
            attempts = TestAttempt.objects.filter(
                test__grade=grade, student__grade=grade, is_completed=True
            ).select_related('student', 'test').order_by('-percentage')
            
            row = 4
            for a in attempts:
                s.cell(row=row, column=1, value=a.student.get_full_name() or a.student.username)
                s.cell(row=row, column=2, value=a.test.subject)
                s.cell(row=row, column=3, value=a.test.title)
                s.cell(row=row, column=4, value=a.score)
                s.cell(row=row, column=5, value=f"{a.percentage:.1f}%")
                
                if a.percentage >= 80:
                    st, f = "A'lo", "C6EFCE"
                elif a.percentage >= 60:
                    st, f = "Yaxshi", "FFEB9C"
                else:
                    st, f = "Qoniqarsiz", "FFC7CE"
                
                s.cell(row=row, column=6, value=st)
                s.cell(row=row, column=7, value=a.finished_at.strftime('%d.%m.%Y') if a.finished_at else '')
                
                for col in range(1, 8):
                    s.cell(row=row, column=col).fill = PatternFill(start_color=f, end_color=f, fill_type="solid")
                row += 1
            
            s.column_dimensions['A'].width = 25
            s.column_dimensions['B'].width = 15
            s.column_dimensions['C'].width = 30
            s.column_dimensions['D'].width = 10
            s.column_dimensions['E'].width = 10
            s.column_dimensions['F'].width = 12
            s.column_dimensions['G'].width = 15
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sinflararo_natijalar.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
