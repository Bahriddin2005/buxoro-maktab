"""
Umumiy test natijalari uchun views
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Max
from .models import TestAttempt


@login_required
def test_api_view(request):
    """API test sahifasi"""
    return render(request, 'tests_app/test_api.html')


@login_required
def student_overall_results_view(request):
    """O'quvchining barcha testlar bo'yicha umumiy natijalari"""
    # Admin va teacher ham o'z natijalarini ko'rishi mumkin
    if request.user.role not in ['student', 'admin', 'teacher']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.headers.get('Accept') == 'application/json' or request.GET.get('format') == 'json':
        try:
            # O'quvchining BARCHA tugallangan testlari
            # Har bir test uchun eng so'nggi attempt'ni olish (bir test bir marta ko'rsatiladi)
            latest_attempt_ids_list = list(TestAttempt.objects.filter(
                student=request.user,
                is_completed=True
            ).values('test').annotate(
                latest_attempt_id=Max('id')
            ).values_list('latest_attempt_id', flat=True))
            
            if not latest_attempt_ids_list:
                attempts = []
            else:
                # Eng so'nggi attempt'larni olish
                attempts = list(TestAttempt.objects.filter(
                    id__in=latest_attempt_ids_list
                ).select_related('test').prefetch_related('result').order_by('-finished_at'))
            
            print(f"=== Overall Results Debug ===")
            print(f"User: {request.user.username}")
            print(f"Found {len(attempts)} unique tests (latest attempts)")
            print(f"Attempt IDs: {latest_attempt_ids_list}")
            
            # Barcha attempt'larni ham yuklab olish (Barcha natijalar uchun)
            all_attempts_for_stats = list(TestAttempt.objects.filter(
                student=request.user,
                is_completed=True
            ).select_related('test').prefetch_related('result').order_by('-finished_at'))
            
            print(f"Total completed attempts: {len(all_attempts_for_stats)}")
            
            # Har bir testning ma'lumotlarini chiqarish
            for attempt in attempts:
                print(f"  - Test: {attempt.test.title}, Score: {attempt.score}/{attempt.total_points}, Percentage: {attempt.percentage}%")
            
            # Barcha attempt'larni ham qaytarish uchun list (Barcha natijalar jadvali uchun)
            all_attempts_list = all_attempts_for_stats
        except Exception as e:
            print(f"Error fetching attempts: {str(e)}")
            import traceback
            traceback.print_exc()
            # Exception bo'lsa ham bo'sh listlarni qaytarish
            return JsonResponse({
                'total_tests': 0,
                'average_score': 0,
                'average_percentage': 0,
                'total_points_earned': 0,
                'total_points_possible': 0,
                'overall_grade': 'N/A',
                'total_correct_answers': 0,
                'total_incorrect_answers': 0,
                'total_unanswered': 0,
                'total_questions': 0,
                'tests': [],
                'all_results': [],  # Bo'sh list qaytarish
                'grade_distribution': {
                    'excellent': 0,
                    'good': 0,
                    'average': 0,
                    'poor': 0
                },
                'subject_stats': {},
                'highest_score': 0,
                'lowest_score': 0
            }, json_dumps_params={'ensure_ascii': False})
        
        # Umumiy statistika
        total_tests = len(attempts)
        
        # Agar hech qanday test bo'lmasa, faqat all_results bo'sh bo'lsa qaytarish
        if total_tests == 0 and len(all_attempts_for_stats) == 0:
            return JsonResponse({
                'total_tests': 0,
                'average_score': 0,
                'average_percentage': 0,
                'total_points_earned': 0,
                'total_points_possible': 0,
                'overall_grade': 'N/A',
                'total_correct_answers': 0,
                'total_incorrect_answers': 0,
                'total_unanswered': 0,
                'total_questions': 0,
                'tests': [],
                'grade_distribution': {
                    'excellent': 0,
                    'good': 0,
                    'average': 0,
                    'poor': 0
                },
                'subject_stats': {},
                'highest_score': 0,
                'lowest_score': 0,
                'all_results': []  # Bo'sh list qaytarish
            }, json_dumps_params={'ensure_ascii': False})
        
        # Hisoblashlar
        try:
            total_score = sum((attempt.score or 0) for attempt in attempts)
            total_points = sum((attempt.total_points or 0) for attempt in attempts)
            average_percentage = (total_score / total_points * 100) if total_points > 0 else 0
        except Exception as e:
            print(f"Error calculating totals: {str(e)}")
            total_score = 0
            total_points = 0
            average_percentage = 0
        
        # Baholar bo'yicha taqsimlash (yangi tizim: 81-100: A'lo, 51-80: Yaxshi, 31-50: Qoniqarli, 0-30: Qoniqarsiz)
        grade_distribution = {
            'excellent': len([a for a in attempts if (a.percentage or 0) >= 81]),
            'good': len([a for a in attempts if 51 <= (a.percentage or 0) < 81]),
            'average': len([a for a in attempts if 31 <= (a.percentage or 0) < 51]),
            'poor': len([a for a in attempts if (a.percentage or 0) < 31])
        }
        
        # Fanlar bo'yicha statistika
        subject_stats = {}
        for attempt in attempts:
            subject = attempt.test.subject
            if subject not in subject_stats:
                subject_stats[subject] = {
                    'test_count': 0,
                    'total_score': 0,
                    'total_possible': 0,
                    'tests': []
                }
            
            subject_stats[subject]['test_count'] += 1
            subject_stats[subject]['total_score'] += attempt.score or 0
            subject_stats[subject]['total_possible'] += attempt.total_points or 0
            subject_stats[subject]['tests'].append(attempt.percentage or 0)
        
        # Har bir fan uchun o'rtacha foiz va baho'ni hisoblash
        for subject, stats in subject_stats.items():
            avg_percentage = (stats['total_score'] / stats['total_possible'] * 100) if stats['total_possible'] > 0 else 0
            stats['average_percentage'] = round(avg_percentage, 1)
            
            # Baho'ni aniqlash (yangi tizim: 81-100: A'lo, 51-80: Yaxshi, 31-50: Qoniqarli, 0-30: Qoniqarsiz)
            if avg_percentage >= 81:
                stats['grade'] = "A'lo"
            elif avg_percentage >= 51:
                stats['grade'] = "Yaxshi"
            elif avg_percentage >= 31:
                stats['grade'] = "Qoniqarli"
            else:
                stats['grade'] = "Qoniqarsiz"
            
            # tests listini o'chirish (kerak emas)
            del stats['tests']
        
        # Umumiy baho (yangi tizim: 81-100: A'lo, 51-80: Yaxshi, 31-50: Qoniqarli, 0-30: Qoniqarsiz)
        if average_percentage >= 81:
            overall_grade = "A'lo"
        elif average_percentage >= 51:
            overall_grade = "Yaxshi"
        elif average_percentage >= 31:
            overall_grade = "Qoniqarli"
        else:
            overall_grade = "Qoniqarsiz"
        
        # Har bir testning ma'lumotlari (har bir test uchun eng so'nggi attempt)
        tests_data = []
        for attempt in attempts:
            # Baho'ni hisoblash
            percentage = attempt.percentage or 0
            if percentage >= 81:
                result_grade = "A'lo"
            elif percentage >= 51:
                result_grade = "Yaxshi"
            elif percentage >= 31:
                result_grade = "Qoniqarli"
            else:
                result_grade = "Qoniqarsiz"
            
            # TestResult dan ma'lumotlarni olish
            try:
                test_result = getattr(attempt, 'result', None)
                if test_result:
                    correct_answers = test_result.correct_answers or 0
                    incorrect_answers = test_result.incorrect_answers or 0
                    unanswered = test_result.unanswered or 0
                else:
                    correct_answers = 0
                    incorrect_answers = 0
                    unanswered = 0
            except Exception as e:
                print(f"Error getting TestResult for attempt {attempt.id}: {str(e)}")
                correct_answers = 0
                incorrect_answers = 0
                unanswered = 0
            
            tests_data.append({
                'test_id': attempt.test.id,
                'test_title': attempt.test.title,
                'subject': attempt.test.subject,
                'grade': attempt.test.grade,
                'score': attempt.score or 0,
                'total_points': attempt.total_points or 0,
                'percentage': round(percentage, 1),
                'result_grade': result_grade,
                'correct_answers': correct_answers,
                'incorrect_answers': incorrect_answers,
                'unanswered': unanswered,
                'time_taken': str(attempt.time_taken) if attempt.time_taken else '0:00:00',
                'finished_at': attempt.finished_at.isoformat() if attempt.finished_at else ''
            })
        
        # Barcha natijalar (har bir test uchun eng so'nggi attempt emas, BARCHA)
        all_results_data = []
        print(f"Processing {len(all_attempts_for_stats)} attempts for all_results")
        for attempt in all_attempts_for_stats:
            # Baho'ni hisoblash
            percentage = attempt.percentage or 0
            if percentage >= 81:
                result_grade = "A'lo"
            elif percentage >= 51:
                result_grade = "Yaxshi"
            elif percentage >= 31:
                result_grade = "Qoniqarli"
            else:
                result_grade = "Qoniqarsiz"
            
            all_results_data.append({
                'test_id': attempt.test.id,
                'test_title': attempt.test.title,
                'subject': attempt.test.subject,
                'grade': attempt.test.grade,
                'score': attempt.score or 0,
                'total_points': attempt.total_points or 0,
                'percentage': round(percentage, 1),
                'result_grade': result_grade,
                'finished_at': attempt.finished_at.isoformat() if attempt.finished_at else ''
            })
        
        print(f"Created {len(all_results_data)} all_results entries")
        
        # Umumiy ballni hisoblash
        total_correct = sum(getattr(getattr(attempt, 'result', None), 'correct_answers', 0) or 0 for attempt in attempts)
        total_incorrect = sum(getattr(getattr(attempt, 'result', None), 'incorrect_answers', 0) or 0 for attempt in attempts)
        total_unanswered = sum(getattr(getattr(attempt, 'result', None), 'unanswered', 0) or 0 for attempt in attempts)
        total_questions = total_correct + total_incorrect + total_unanswered
        
        try:
            return JsonResponse({
                'total_tests': total_tests,
                'average_score': round(total_score / total_tests, 1) if total_tests > 0 else 0,
                'average_percentage': round(average_percentage, 1),
                'total_points_earned': total_score,
                'total_points_possible': total_points,
                'overall_grade': overall_grade,
                'total_correct_answers': total_correct,
                'total_incorrect_answers': total_incorrect,
                'total_unanswered': total_unanswered,
                'total_questions': total_questions,
                'total_questions_overall': total_questions,  # Alias for frontend
                'tests': tests_data,
                'all_results': all_results_data,  # Barcha natijalar
                'grade_distribution': grade_distribution,
                'subject_stats': subject_stats,
                'highest_score': round(max((attempt.percentage or 0) for attempt in attempts), 1) if attempts else 0,
                'lowest_score': round(min((attempt.percentage or 0) for attempt in attempts), 1) if attempts else 0
            }, json_dumps_params={'ensure_ascii': False})
        except Exception as final_error:
            print(f"Error in final JsonResponse: {str(final_error)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'error': f'Ma\'lumotlarni qaytarishda xatolik: {str(final_error)}',
                'total_tests': 0,
                'average_percentage': 0,
                'tests': [],
                'all_results': [],
                'grade_distribution': {'excellent': 0, 'good': 0, 'average': 0, 'poor': 0},
                'subject_stats': {}
            }, json_dumps_params={'ensure_ascii': False}, status=500)
    
    return render(request, 'tests_app/overall_results.html')


@login_required
def student_export_results_view(request):
    """O'quvchining barcha test natijalarini Excel formatida export qilish"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.db.models import Max
        
        # O'quvchining barcha tugallangan testlari
        latest_attempts = TestAttempt.objects.filter(
            student=request.user,
            is_completed=True
        ).values('test').annotate(
            latest_attempt_id=Max('id')
        ).values_list('latest_attempt_id', flat=True)
        
        attempts = TestAttempt.objects.filter(
            id__in=latest_attempts
        ).select_related('test').order_by('-finished_at')
        
        if attempts.count() == 0:
            return JsonResponse({'error': 'Test natijalari topilmadi'}, status=404)
        
        # Yangi workbook yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Mening Natijalarim"
        
        # O'quvchi ma'lumotlari
        student_name = f"{request.user.first_name} {request.user.last_name}"
        if not student_name.strip():
            student_name = request.user.username
        
        ws['A1'] = "TEST NATIJALARI"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:G1')
        
        ws['A2'] = f"O'quvchi: {student_name}"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:G2')
        
        ws['A3'] = f"Sinf: {request.user.grade}"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A3:G3')
        
        # Umumiy statistika
        total_score = sum(attempt.score for attempt in attempts)
        total_points = sum(attempt.total_points for attempt in attempts)
        average_percentage = (total_score / total_points * 100) if total_points > 0 else 0
        
        if average_percentage >= 81:
            overall_grade = "A'lo"
        elif average_percentage >= 61:
            overall_grade = "Yaxshi"
        elif average_percentage >= 31:
            overall_grade = "Qoniqarli"
        else:
            overall_grade = "Qoniqarsiz"
        
        ws['A5'] = "UMUMIY STATISTIKA"
        ws['A5'].font = Font(bold=True, size=12)
        ws.merge_cells('A5:G5')
        
        ws['A6'] = f"Jami testlar: {attempts.count()}"
        ws['C6'] = f"O'rtacha natija: {average_percentage:.1f}%"
        ws['E6'] = f"Umumiy baho: {overall_grade}"
        
        # Header
        headers = ['№', 'Test nomi', 'Fan', 'Ball', 'Foiz', 'Baho', 'Sana']
        row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Test natijalari
        row = 9
        for i, attempt in enumerate(attempts, 1):
            percentage = attempt.percentage or 0
            
            # Baho'ni hisoblash
            if percentage >= 81:
                result_grade = "A'lo"
                fill_color = "C6EFCE"  # Yashil
            elif percentage >= 61:
                result_grade = "Yaxshi"
                fill_color = "FFEB9C"  # Sariq
            elif percentage >= 31:
                result_grade = "Qoniqarli"
                fill_color = "BDD7EE"  # Ko'k
            else:
                result_grade = "Qoniqarsiz"
                fill_color = "FFC7CE"  # Qizil
            
            # Ma'lumotlarni qo'shish
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=attempt.test.title)
            ws.cell(row=row, column=3, value=attempt.test.subject)
            ws.cell(row=row, column=4, value=f"{attempt.score}/{attempt.total_points}")
            ws.cell(row=row, column=5, value=f"{percentage:.1f}%")
            ws.cell(row=row, column=6, value=result_grade)
            ws.cell(row=row, column=7, value=attempt.finished_at.strftime('%d.%m.%Y') if attempt.finished_at else '')
            
            # Rang berish va border
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        # Fanlar bo'yicha statistika
        row += 2
        ws[f'A{row}'] = "FANLAR BO'YICHA NATIJALAR"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        # Fanlar bo'yicha statistika hisoblash
        subject_stats = {}
        for attempt in attempts:
            subject = attempt.test.subject
            if subject not in subject_stats:
                subject_stats[subject] = {
                    'test_count': 0,
                    'total_score': 0,
                    'total_possible': 0
                }
            
            subject_stats[subject]['test_count'] += 1
            subject_stats[subject]['total_score'] += attempt.score
            subject_stats[subject]['total_possible'] += attempt.total_points
        
        # Header
        ws[f'A{row}'] = "Fan"
        ws[f'B{row}'] = "Testlar soni"
        ws[f'C{row}'] = "Ball"
        ws[f'D{row}'] = "Foiz"
        ws[f'E{row}'] = "Baho"
        
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
        
        # Har bir fan uchun
        for subject, stats in subject_stats.items():
            avg_percentage = (stats['total_score'] / stats['total_possible'] * 100) if stats['total_possible'] > 0 else 0
            
            if avg_percentage >= 81:
                grade = "A'lo"
                fill_color = "C6EFCE"
            elif avg_percentage >= 61:
                grade = "Yaxshi"
                fill_color = "FFEB9C"
            elif avg_percentage >= 31:
                grade = "Qoniqarli"
                fill_color = "BDD7EE"
            else:
                grade = "Qoniqarsiz"
                fill_color = "FFC7CE"
            
            ws[f'A{row}'] = subject
            ws[f'B{row}'] = stats['test_count']
            ws[f'C{row}'] = f"{stats['total_score']}/{stats['total_possible']}"
            ws[f'D{row}'] = f"{avg_percentage:.1f}%"
            ws[f'E{row}'] = grade
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row += 1
        
        # Baholar bo'yicha statistika
        row += 2
        ws[f'A{row}'] = "BAHOLAR BO'YICHA TAQSIMLASH"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        excellent_count = attempts.filter(percentage__gte=81).count()
        good_count = attempts.filter(percentage__gte=61, percentage__lt=81).count()
        average_count = attempts.filter(percentage__gte=31, percentage__lt=61).count()
        poor_count = attempts.filter(percentage__lt=31).count()
        
        ws[f'A{row}'] = f"A'lo (81%+): {excellent_count} ta"
        ws[f'C{row}'] = f"Yaxshi (61-80%): {good_count} ta"
        ws[f'E{row}'] = f"Qoniqarli (31-60%): {average_count} ta"
        ws[f'G{row}'] = f"Qoniqarsiz (0-30%): {poor_count} ta"
        
        # Ustun kengliklarini sozlash
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        # Response yaratish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{student_name}_test_natijalari.xlsx"'
        
        # Excel faylni saqlash
        wb.save(response)
        
        return response
        
    except Exception as e:
        print(f"Student export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Export xatolik yuz berdi: {str(e)}'}, status=500)

