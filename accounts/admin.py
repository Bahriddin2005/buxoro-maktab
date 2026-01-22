from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.utils import timezone
from django import forms
from django.contrib.auth.forms import UserChangeForm
from django.utils.html import format_html

# Faqat mavjud modellarni import qilamiz
try:
    from .models import User, VerificationRequest
except ImportError:
    User = None
    VerificationRequest = None

# Agar User modeli mavjud bo'lsa va custom fieldlarga ega bo'lsa
if User and hasattr(User, 'role'):
    
    class CustomUserChangeForm(UserChangeForm):
        class Meta(UserChangeForm.Meta):
            model = User
            fields = '__all__'
    
    class CustomUserAdmin(UserAdmin):
        model = User
        form = CustomUserChangeForm
        
        list_display = ['username', 'full_name', 'email', 'role_badge', 'grade_badge', 'is_verified_icon', 'student_id', 'is_active_icon']
        list_filter = ['role', 'is_verified', 'school_email_verified', 'is_active', 'grade']
        search_fields = ['username', 'email', 'student_id', 'first_name', 'last_name']
        list_per_page = 30
        
        def full_name(self, obj):
            return f"{obj.first_name} {obj.last_name}" if obj.first_name else obj.username
        full_name.short_description = 'Ism Familiya'
        
        def role_badge(self, obj):
            colors = {
                'admin': '#dc3545',
                'teacher': '#007bff',
                'student': '#28a745'
            }
            icons = {
                'admin': '👑',
                'teacher': '👨‍🏫',
                'student': '👨‍🎓'
            }
            return format_html(
                '<span style="background-color: {}; color: white; padding: 4px 12px; border-radius: 6px; font-weight: bold;">{} {}</span>',
                colors.get(obj.role, '#6c757d'),
                icons.get(obj.role, ''),
                obj.get_role_display() if hasattr(obj.get_role_display, '__call__') else obj.role
            )
        role_badge.short_description = 'Rol'
        
        def grade_badge(self, obj):
            if obj.grade:
                return format_html(
                    '<span style="background-color: #6f42c1; color: white; padding: 4px 10px; border-radius: 6px; font-weight: bold;">{}-sinf</span>',
                    obj.grade
                )
            return '-'
        grade_badge.short_description = 'Sinf'
        
        def is_verified_icon(self, obj):
            if obj.is_verified:
                return format_html('<span style="color: green; font-size: 18px;" title="Tasdiqlangan">✓</span>')
            return format_html('<span style="color: red; font-size: 18px;" title="Tasdiqlanmagan">✗</span>')
        is_verified_icon.short_description = 'Tasdiqlangan'
        
        def is_active_icon(self, obj):
            if obj.is_active:
                return format_html('<span style="color: green; font-size: 18px;" title="Aktiv">✓</span>')
            return format_html('<span style="color: red; font-size: 18px;" title="Faol emas">✗</span>')
        is_active_icon.short_description = 'Aktiv'
        
        # To'liq qayta aniqlangan fieldsets
        fieldsets = (
            (None, {'fields': ('username', 'password')}),
            ('Personal info', {'fields': ('first_name', 'last_name', 'email')}),
            ('Permissions', {
                'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
            }),
            ('Important dates', {'fields': ('last_login',)}),  # Faqat last_login
            ('School Info', {
                'fields': ('role', 'student_id', 'is_verified', 'school_email_verified', 
                          'phone_number', 'class_name', 'grade', 'subject')
            }),
        )
        
        add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': ('username', 'password1', 'password2'),
            }),
            ('Personal info', {'fields': ('first_name', 'last_name', 'email')}),
            ('School Info', {
                'fields': ('role', 'student_id', 'phone_number', 'class_name', 'grade', 'subject')
            }),
        )
        
        filter_horizontal = ('groups', 'user_permissions',)
        
        # Export actions
        def export_teachers_credentials(self, request, queryset):
            """Export teachers credentials to Excel"""
            from .management.commands.export_user_credentials import Command
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.utils import timezone
            from django.http import HttpResponse
            
            teachers = User.objects.filter(role='teacher', is_verified=True).order_by('first_name', 'last_name')
            
            if not teachers.exists():
                self.message_user(request, 'O\'qituvchilar topilmadi!', level='warning')
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "O'qituvchilar"
            
            # Title
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = "O'QITUVCHILAR LOGIN VA PAROLLARI"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # Info
            ws.merge_cells('A2:E2')
            info_cell = ws['A2']
            info_cell.value = f"Export qilingan sana: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Jami: {teachers.count()} ta"
            info_cell.font = Font(size=10, italic=True)
            info_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 20
            
            # Headers
            headers = ['№', 'Ism', 'Familiya', 'Login (Username)', 'Email', 'Fan']
            header_row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[header_row].height = 25
            
            # Data
            for idx, user in enumerate(teachers, 1):
                row = header_row + idx
                data = [
                    idx,
                    user.first_name or '',
                    user.last_name or '',
                    user.username,
                    user.email,
                    user.subject or '',
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 20
            
            # Save to response
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="oqituvchilar_login_parol_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            self.message_user(request, f'{teachers.count()} ta o\'qituvchi ma\'lumotlari yuklab olindi!')
            return response
        
        def export_students_credentials(self, request, queryset):
            """Export students credentials to Excel"""
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.utils import timezone
            from django.http import HttpResponse
            
            students = User.objects.filter(role='student', is_verified=True).order_by('grade', 'class_name', 'first_name', 'last_name')
            
            if not students.exists():
                self.message_user(request, 'O\'quvchilar topilmadi!', level='warning')
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "O'quvchilar"
            
            # Title
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = "O'QUVCHILAR LOGIN VA PAROLLARI"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # Info
            ws.merge_cells('A2:G2')
            info_cell = ws['A2']
            info_cell.value = f"Export qilingan sana: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Jami: {students.count()} ta"
            info_cell.font = Font(size=10, italic=True)
            info_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 20
            
            # Headers
            headers = ['№', 'Ism', 'Familiya', 'Login (Username)', 'Email', 'Sinf', 'Sinif']
            header_row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[header_row].height = 25
            
            # Data
            for idx, user in enumerate(students, 1):
                row = header_row + idx
                data = [
                    idx,
                    user.first_name or '',
                    user.last_name or '',
                    user.username,
                    user.email,
                    user.grade or '',
                    user.class_name or '',
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15
            
            # Save to response
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="oquvchilar_login_parol_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            self.message_user(request, f'{students.count()} ta o\'quvchi ma\'lumotlari yuklab olindi!')
            return response
        
        export_teachers_credentials.short_description = "📥 O'qituvchilar login va parollarini yuklab olish"
        export_students_credentials.short_description = "📥 O'quvchilar login va parollarini yuklab olish"
        
        actions = [export_teachers_credentials, export_students_credentials]
    
    admin.site.register(User, CustomUserAdmin)

# VerificationRequest admin
if VerificationRequest:
    @admin.register(VerificationRequest)
    class VerificationRequestAdmin(admin.ModelAdmin):
        list_display = ['user_info', 'requested_at', 'status_badge', 'processed_by', 'processed_at']
        list_filter = ['is_approved', 'requested_at']
        search_fields = ['user__username', 'user__email', 'user__first_name', 'user__last_name']
        readonly_fields = ['requested_at']
        list_per_page = 20
        date_hierarchy = 'requested_at'
        
        def user_info(self, obj):
            user = obj.user
            name = f"{user.first_name} {user.last_name}" if user.first_name else user.username
            role_colors = {
                'admin': '#dc3545',
                'teacher': '#007bff',
                'student': '#28a745'
            }
            return format_html(
                '{} <span style="background-color: {}; color: white; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: bold;">{}</span>',
                name,
                role_colors.get(user.role, '#6c757d'),
                user.get_role_display() if hasattr(user.get_role_display, '__call__') else user.role
            )
        user_info.short_description = 'Foydalanuvchi'
        
        def status_badge(self, obj):
            if obj.is_approved is None:
                return format_html('<span style="background-color: #ffc107; color: white; padding: 4px 12px; border-radius: 6px; font-weight: bold;">⏳ Kutilmoqda</span>')
            elif obj.is_approved:
                return format_html('<span style="background-color: #28a745; color: white; padding: 4px 12px; border-radius: 6px; font-weight: bold;">✓ Tasdiqlangan</span>')
            else:
                return format_html('<span style="background-color: #dc3545; color: white; padding: 4px 12px; border-radius: 6px; font-weight: bold;">✗ Rad etilgan</span>')
        status_badge.short_description = 'Holat'
        
        def approve_request(self, request, queryset):
            for verification_request in queryset:
                verification_request.user.is_verified = True
                verification_request.user.save()
                verification_request.is_approved = True
                verification_request.processed_by = request.user
                verification_request.processed_at = timezone.now()
                verification_request.save()
            self.message_user(request, f'{queryset.count()} requests approved.')
        
        def reject_request(self, request, queryset):
            for verification_request in queryset:
                verification_request.is_approved = False
                verification_request.processed_by = request.user
                verification_request.processed_at = timezone.now()
                verification_request.save()
            self.message_user(request, f'{queryset.count()} requests rejected.')
        
        approve_request.short_description = "Approve selected requests"
        reject_request.short_description = "Reject selected requests"
        actions = [approve_request, reject_request]