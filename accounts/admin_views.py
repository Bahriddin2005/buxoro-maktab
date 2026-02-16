"""
Admin panel uchun alohida view'lar
"""
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from functools import wraps
from io import BytesIO
import secrets
import string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from accounts.models import User


def admin_required(view_func):
    """Faqat admin roli bo'lgan foydalanuvchilar uchun decorator"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return HttpResponseForbidden('Tizimga kiring!')
        
        # Superuser yoki admin rolida bo'lishi kerak
        is_admin = (
            request.user.is_superuser or 
            (hasattr(request.user, 'role') and request.user.role == 'admin')
        )
        
        if not is_admin:
            return HttpResponseForbidden('Bu funksiya faqat adminlar uchun!')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def generate_secure_password(length=12):
    """Xavfsiz tasodifiy parol generatsiya qilish"""
    # Parol tarkibi: harflar, raqamlar, maxsus belgilar
    alphabet = string.ascii_letters + string.digits
    # Kamida bitta katta harf, bitta kichik harf, bitta raqam bo'lishini ta'minlash
    password = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ]
    # Qolgan belgilarni qo'shish
    password += [secrets.choice(alphabet) for _ in range(length - 3)]
    # Aralashtirish
    secrets.SystemRandom().shuffle(password)
    return ''.join(password)


@admin_required
def export_teachers_credentials_view(request):
    """
    O'qituvchilar login, email va JORIY PAROLLARINI yuklab olish (FAQAT ADMIN UCHUN)
    Faqat mavjud (saqlangan) parollar ko'rsatiladi
    """
    teachers = User.objects.filter(role='teacher', is_verified=True).order_by('first_name', 'last_name')
    
    if not teachers.exists():
        return HttpResponse('O\'qituvchilar topilmadi!', status=404)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "O'qituvchilar"
    
    # Thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers: Login, Email, Parol
    headers = ['№', 'Login (Username)', 'Email', 'Parol']
    merge_range = 'A1:D1'
    merge_range_info = 'A2:D2'
    
    # Title
    ws.merge_cells(merge_range)
    title_cell = ws['A1']
    title_cell.value = "O'QITUVCHILAR LOGIN, EMAIL VA PAROLLARI"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Info
    ws.merge_cells(merge_range_info)
    info_cell = ws['A2']
    info_text = f"Export qilingan sana: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Jami: {teachers.count()} ta"
    info_cell.value = info_text
    info_cell.font = Font(size=10, italic=True)
    info_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Headers
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 25
    
    # Data
    for idx, user in enumerate(teachers, 1):
        row = header_row + idx
        
        # Parol: o'zgargan, yangi qo'shilgan yoki parolsiz (avtomatik yangi parol)
        password = None
        try:
            if hasattr(user, 'temporary_password') and user.temporary_password:
                password = user.temporary_password
        except (AttributeError, Exception):
            pass
        
        # Parolsiz bo'lsa - avtomatik yangi parol yaratib saqlash
        if not password:
            password = generate_secure_password()
            user.set_password(password)
            user.temporary_password = password
            user.save()
        
        data = [
            idx,
            user.username,
            user.email,
            password,
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    
    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"oqituvchilar_login_parol_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@admin_required
def export_students_credentials_view(request):
    """
    O'quvchilar login, email va JORIY PAROLLARINI yuklab olish (FAQAT ADMIN UCHUN)
    Faqat mavjud (saqlangan) parollar ko'rsatiladi
    """
    students = User.objects.filter(role='student', is_verified=True).order_by('grade', 'class_name', 'first_name', 'last_name')
    
    if not students.exists():
        return HttpResponse('O\'quvchilar topilmadi!', status=404)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"
    
    # Thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers: Login, Email, Parol
    headers = ['№', 'Login (Username)', 'Email', 'Parol']
    merge_range = 'A1:D1'
    merge_range_info = 'A2:D2'
    
    # Title
    ws.merge_cells(merge_range)
    title_cell = ws['A1']
    title_cell.value = "O'QUVCHILAR LOGIN, EMAIL VA PAROLLARI"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Info
    ws.merge_cells(merge_range_info)
    info_cell = ws['A2']
    info_text = f"Export qilingan sana: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Jami: {students.count()} ta"
    info_cell.value = info_text
    info_cell.font = Font(size=10, italic=True)
    info_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Headers
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 25
    
    # Data
    for idx, user in enumerate(students, 1):
        row = header_row + idx
        
        # Parol: o'zgargan, yangi qo'shilgan yoki parolsiz (avtomatik yangi parol)
        password = None
        try:
            if hasattr(user, 'temporary_password') and user.temporary_password:
                password = user.temporary_password
        except (AttributeError, Exception):
            pass
        
        # Parolsiz bo'lsa - avtomatik yangi parol yaratib saqlash
        if not password:
            password = generate_secure_password()
            user.set_password(password)
            user.temporary_password = password
            user.save()
        
        data = [
            idx,
            user.username,
            user.email,
            password,
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    
    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"oquvchilar_login_parol_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@admin_required
def download_import_template_view(request):
    """Import uchun Excel shablonini yuklab olish"""
    wb = Workbook()

    # 1-varaq: O'quvchilar (namuna va sarlavhalar)
    ws = wb.active
    ws.title = "O'quvchilar"
    headers = ['Login', 'Parol', 'Ism', 'Familiya', 'Sinf']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    # Namuna qatorlar
    samples = [
        ['ali_valiyev', 'Parol123', 'Ali', 'Valiyev', 5],
        ['malika_karimova', 'Parol456', 'Malika', 'Karimova', 5],
        ['sardor_ismoilov', 'Parol789', 'Sardor', 'Ismoilov', 6],
    ]
    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 8

    # 2-varaq: Qo'llanma
    ws2 = wb.create_sheet(title="Qo'llanma")
    instructions = [
        ["O'QUVCHILARNI WEB-SAYTGA QANDAY TO'G'RI QO'SHISH"],
        [""],
        ["1-QADAM: Shablonni tushiring"],
        ["   • «Shablonni yuklab olish» tugmasini bosing"],
        ["   • O'quvchilar varag'idagi namuna qatorlarni o'chiring (2–4 qatorlar)"],
        [""],
        ["2-QADAM: O'quvchilar ma'lumotlarini kiriting"],
        ["   • Login — tizimga kirish uchun (harflar, raqamlar, _ belgisi)"],
        ["   • Parol — har bir o'quvchining paroli"],
        ["   • Ism, Familiya — ixtiyoriy (bo'sh qoldirish mumkin)"],
        ["   • Sinf — 1 dan 11 gacha (agar bo'sh bo'lsa, import sahifasida tanlangan sinf qo'llanadi)"],
        [""],
        ["3-QADAM: Faylni saqlang va import qiling"],
        ["   • Excel faylni .xlsx formatida saqlang"],
        ["   • Saytda «O'quvchilarni Excel'dan Import» sahifasiga o'ting"],
        ["   • Sinfni tanlang (agar Excel'da Sinf ustuni bo'lmasa)"],
        ["   • Faylni yuklang va «Import qilish» tugmasini bosing"],
        [""],
        ["Muhim:"],
        ["   • Birinchi qator sarlavha bo'lishi SHART (Login, Parol, Ism, Familiya, Sinf)"],
        ["   • Login va Parol ustunlari bo'sh bo'lmasin"],
        ["   • Mavjud login qayta yozilsa — parol va sinf yangilanadi"],
        ["   • O'quvchilar qo'shilgach — o'z sinflaridagi testlarni yecha olishadi"],
    ]
    for row_idx, row_text in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=row_text[0])
        cell = ws2.cell(row=row_idx, column=1)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_text[0].startswith(("1-QADAM", "2-QADAM", "3-QADAM", "Muhim:")):
            cell.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 70

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="oquvchilar_import_shabloni.xlsx"'
    return response


@login_required
@admin_required
@require_http_methods(["GET", "POST"])
def import_students_from_excel_view(request):
    """
    Excel fayldan o'quvchilarni import qilish.
    Excel format: Login, Parol, (ixtiyoriy: Ism, Familiya, Sinf)
    Yoki: 1-qator sarlavha, 2-qatordan ma'lumotlar.
    """
    if request.method == 'GET':
        return render(request, 'admin/import_students.html')

    # POST - fayl qabul qilish
    excel_file = request.FILES.get('excel_file')
    grade_override = request.POST.get('grade')  # Barcha qatorlar uchun sinf (ixtiyoriy)

    if not excel_file:
        messages.error(request, "Excel fayl tanlanmadi!")
        return redirect('accounts:import_students')

    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, "Faqat .xlsx formatidagi fayllar qabul qilinadi!")
        return redirect('accounts:import_students')

    try:
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"Excel faylni o'qishda xatolik: {str(e)}")
        return redirect('accounts:import_students')

    # Ustun nomlarini topish (birinchisi Login, Parol, Ism, Familiya, Sinf)
    COL_ALIASES = {
        'login': ['login', 'username', 'login (username)', 'логин'],
        'parol': ['parol', 'password', 'пароль'],
        'ism': ['ism', 'first_name', 'first name', 'имя', 'name'],
        'familiya': ['familiya', 'last_name', 'last name', 'фамилия'],
        'sinf': ['sinf', 'grade', 'класс', 'class'],
    }

    def find_column_index(row, key):
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            val = str(cell_val).strip().lower()
            for alias in COL_ALIASES.get(key, []):
                if alias in val or val in alias:
                    return col_idx
        return -1

    created_count = 0
    updated_count = 0
    errors = []

    try:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "Excel faylda ma'lumot yo'q!")
            return redirect('accounts:import_students')

        # Birinchi qator - sarlavha yoki ma'lumot
        first_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        has_header = any(
            'login' in r or 'parol' in r or 'username' in r or 'password' in r
            for r in first_row
        )
        start_row = 1 if has_header else 0

        # Header bo'lsa, ustun indekslarini bir marta hisoblaymiz
        if has_header and len(rows) > 0:
            login_idx = find_column_index(rows[0], 'login')
            parol_idx = find_column_index(rows[0], 'parol')
            ism_idx = find_column_index(rows[0], 'ism')
            fam_idx = find_column_index(rows[0], 'familiya')
            sinf_idx = find_column_index(rows[0], 'sinf')
        else:
            login_idx, parol_idx = 0, 1
            ism_idx, fam_idx = 2, 3
            sinf_idx = 4 if len(rows[0]) > 4 else -1 if rows else -1

        for row_idx, row in enumerate(rows[start_row:], start=start_row + 1):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            login = None
            if 0 <= login_idx < len(row) and row[login_idx] is not None:
                login = str(row[login_idx]).strip()
            parol = None
            if 0 <= parol_idx < len(row) and row[parol_idx] is not None:
                parol = str(row[parol_idx]).strip()

            if not login or not parol:
                errors.append(f"Qator {row_idx}: Login yoki Parol bo'sh")
                continue

            # Username - faqat raqamlar va harflar
            username = ''.join(c for c in login if c.isalnum() or c in '._-')
            if not username:
                username = f"student_{row_idx}"

            first_name = ''
            if 0 <= ism_idx < len(row) and row[ism_idx] is not None:
                first_name = str(row[ism_idx]).strip()
            last_name = ''
            if 0 <= fam_idx < len(row) and row[fam_idx] is not None:
                last_name = str(row[fam_idx]).strip()
            if not first_name and not last_name:
                first_name, last_name = username, ''

            # Sinf: Excel'dan yoki formadan
            grade_val = None
            if 0 <= sinf_idx < len(row) and row[sinf_idx] is not None:
                try:
                    g = str(row[sinf_idx]).strip().replace('-sinf', '').replace('sinf', '').strip()
                    if g.isdigit():
                        grade_val = int(g)
                except (ValueError, TypeError):
                    pass
            if grade_val is None and grade_override:
                try:
                    grade_val = int(grade_override)
                except (ValueError, TypeError):
                    pass

            if grade_val is None or grade_val < 1 or grade_val > 11:
                grade_val = 5  # default

            email = f"{username}@student.buxorobilimdonlar.uz"
            existing = User.objects.filter(username=username).first()
            if existing:
                # Mavjud o'quvchini yangilash
                existing.set_password(parol)
                existing.temporary_password = parol
                existing.grade = grade_val
                existing.first_name = first_name or existing.first_name
                existing.last_name = last_name or existing.last_name
                existing.is_verified = True
                existing.is_active = True
                existing.save()
                updated_count += 1
            else:
                # Yangi o'quvchi - email unikalligini ta'minlash
                base_username = username
                suffix = 0
                while User.objects.filter(username=username).exists() or User.objects.filter(email=email).exists():
                    suffix += 1
                    username = f"{base_username}{suffix}"
                    email = f"{username}@student.buxorobilimdonlar.uz"

                try:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=parol,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    user.role = 'student'
                    user.grade = grade_val
                    user.student_id = username
                    user.is_verified = True
                    user.temporary_password = parol
                    user.save()
                    created_count += 1
                except Exception as e:
                    errors.append(f"Qator {row_idx} ({login}): {str(e)}")

        wb.close()
    except Exception as e:
        messages.error(request, f"Import jarayonida xatolik: {str(e)}")
        return redirect('accounts:import_students')

    if created_count or updated_count:
        msg = f"✅ Muvaffaqiyat! {created_count} ta yangi o'quvchi qo'shildi."
        if updated_count:
            msg += f" {updated_count} ta yangilandi."
        messages.success(request, msg)
    if errors:
        messages.warning(request, f"Ba'zi qatorlarda xatolik ({len(errors)} ta).")
        for err in errors[:5]:
            messages.warning(request, err)
        if len(errors) > 5:
            messages.warning(request, f"... va yana {len(errors) - 5} ta xatolik")

    return redirect('accounts:import_students')
