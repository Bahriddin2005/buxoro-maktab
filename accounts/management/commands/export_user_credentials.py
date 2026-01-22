"""
Django management command to export user credentials (login and passwords)
Separate files for teachers and students
"""
from django.core.management.base import BaseCommand
from accounts.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
import os
from pathlib import Path


class Command(BaseCommand):
    help = 'Export user credentials (login and passwords) to Excel files'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            default='exports',
            help='Directory to save exported files (default: exports)'
        )
        parser.add_argument(
            '--include-passwords',
            action='store_true',
            help='Include passwords in export (default: False)'
        )

    def handle(self, *args, **options):
        output_dir = options['output_dir']
        include_passwords = options['include_passwords']
        
        # Create output directory if it doesn't exist
        BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
        export_path = BASE_DIR / output_dir
        export_path.mkdir(exist_ok=True)
        
        # Get all teachers
        teachers = User.objects.filter(role='teacher', is_verified=True).order_by('first_name', 'last_name')
        
        # Get all students
        students = User.objects.filter(role='student', is_verified=True).order_by('grade', 'class_name', 'first_name', 'last_name')
        
        # Export teachers
        if teachers.exists():
            teachers_file = self.export_users(
                users=teachers,
                user_type='teachers',
                filepath=export_path / f'oqituvchilar_login_parol_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                include_passwords=include_passwords
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ {teachers.count()} ta o\'qituvchi ma\'lumotlari saqlandi: {teachers_file}'
                )
            )
        else:
            self.stdout.write(self.style.WARNING('⚠️ O\'qituvchilar topilmadi!'))
        
        # Export students
        if students.exists():
            students_file = self.export_users(
                users=students,
                user_type='students',
                filepath=export_path / f'oquvchilar_login_parol_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                include_passwords=include_passwords
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ {students.count()} ta o\'quvchi ma\'lumotlari saqlandi: {students_file}'
                )
            )
        else:
            self.stdout.write(self.style.WARNING('⚠️ O\'quvchilar topilmadi!'))
        
        self.stdout.write(
            self.style.SUCCESS(
                f'\n🎉 Export yakunlandi! Fayllar: {export_path}'
            )
        )

    def export_users(self, users, user_type, filepath, include_passwords=False):
        """Export users to Excel file"""
        wb = Workbook()
        ws = wb.active
        
        # Set sheet title
        if user_type == 'teachers':
            ws.title = "O'qituvchilar"
            title = "O'QITUVCHILAR LOGIN VA PAROLLARI"
        else:
            ws.title = "O'quvchilar"
            title = "O'QUVCHILAR LOGIN VA PAROLLARI"
        
        # Title row
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Info row
        ws.merge_cells('A2:F2')
        info_cell = ws['A2']
        info_cell.value = f"Export qilingan sana: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Jami: {users.count()} ta"
        info_cell.font = Font(size=10, italic=True)
        info_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # Headers
        headers = ['№', 'Ism', 'Familiya', 'Login (Username)', 'Email', 'Parol']
        if not include_passwords:
            headers = headers[:-1]  # Remove password column if not included
        
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[header_row].height = 25
        
        # Data rows
        for idx, user in enumerate(users, 1):
            row = header_row + idx
            data = [
                idx,
                user.first_name or '',
                user.last_name or '',
                user.username,
                user.email,
            ]
            
            if include_passwords:
                # Note: Django stores hashed passwords, we can't retrieve original passwords
                # But we can show a note or generate a temporary password
                data.append('(Parol hash qilingan, qayta tiklash kerak)')
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Alternate row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Add additional info for students
        if user_type == 'students':
            # Add grade and class info
            ws.insert_cols(6)  # Insert before password column
            ws.cell(row=header_row, column=6, value='Sinf').font = Font(bold=True, color="FFFFFF")
            ws.cell(row=header_row, column=6).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.cell(row=header_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.insert_cols(7)
            ws.cell(row=header_row, column=7, value='Sinif').font = Font(bold=True, color="FFFFFF")
            ws.cell(row=header_row, column=7).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.cell(row=header_row, column=7).alignment = Alignment(horizontal='center', vertical='center')
            
            # Update data rows for students
            for idx, user in enumerate(users, 1):
                row = header_row + idx
                ws.cell(row=row, column=6, value=user.grade or '')
                ws.cell(row=row, column=7, value=user.class_name or '')
                
                # Alternate row colors
                if row % 2 == 0:
                    for col in range(6, 8):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Add additional info for teachers
        if user_type == 'teachers':
            # Add subject info
            ws.insert_cols(6)
            ws.cell(row=header_row, column=6, value='Fan').font = Font(bold=True, color="FFFFFF")
            ws.cell(row=header_row, column=6).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.cell(row=header_row, column=6).alignment = Alignment(horizontal='center', vertical='center')
            
            # Update data rows for teachers
            for idx, user in enumerate(users, 1):
                row = header_row + idx
                ws.cell(row=row, column=6, value=user.subject or '')
                
                # Alternate row colors
                if row % 2 == 0:
                    ws.cell(row=row, column=6).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Adjust column widths
        column_widths = {
            'A': 8,   # №
            'B': 20,  # Ism
            'C': 20,  # Familiya
            'D': 25,  # Login
            'E': 30,  # Email
            'F': 15,  # Sinf/Fan
            'G': 15,  # Sinif (students only)
            'H': 40,  # Parol (if included)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save file
        wb.save(filepath)
        return str(filepath)
